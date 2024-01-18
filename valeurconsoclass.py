import tkinter
#import tkintermapview
import customtkinter
import os
from PIL import Image
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import subprocess
from datetime import datetime
import time


#il faut une boucle qui affiche la consommation de la traversé précédent et actuelle

customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("blue")  # Themes: blue (default), dark-blue, green

class window_Valeur_de_consommation(customtkinter.CTk):
    def __init__(self,utilisateur):
        super().__init__()
        consommation_précedant=400
        consommation_actuelle=450
        self.geometry(f"{800}x{480}")
        self.utilisateur=utilisateur
          # create CTk window like you do with the Tk window

        self.grid_columnconfigure((0,1,2,3,4), weight=1)
        self.grid_rowconfigure((0, 1, 2,3,4,5,6,7,8,9,10,11,12), weight=1)

        def retour_Window_menu():
            from Menu import window_Menu
            self.destroy()
            window_Menuu=window_Menu(self.utilisateur)
            window_Menuu.mainloop()
            
        def goto_historique(time,totaliseur):
            from Menu import window_Menu
            self.destroy()
            #window_Menu.destroy()
            window_Menu=window_Menu()
            window_Menu.mainloop()

        def button_function():
            print("button pressed")
        carburant1="fuel lourd"
        carburant2="MDO"

        workbook = load_workbook('fichier_excel/consommation.xlsx')
        worksheet = workbook.active
        """
        Row=len(worksheet['A'])
        print(Row,rowmax)
        """
        rowmax=worksheet.max_row
        print(rowmax)
        """
        self.MP1total=tkinter.StringVar()
        self.MP1total.set(str(worksheet.cell(row=rowmax, column=5).value))
        """
        def refresh():
            traversé_quai_found=0
            for i in range(1,rowmax):
                if str(worksheet.cell(row=(rowmax-i), column=2).value)=="traverse":
                    traversé_quai_found+=1
                    print(rowmax)
                    self.MP1depart=str(worksheet.cell(row=(rowmax-i)+1, column=5).value)

                    self.MP2depart=str(worksheet.cell(row=(rowmax-i)+1, column=8).value)

                    self.MP3depart=str(worksheet.cell(row=(rowmax-i)+1, column=11).value)

                    self.MP4depart=str(worksheet.cell(row=(rowmax-i)+1, column=14).value)
                    
                    self.GE1depart=str(worksheet.cell(row=(rowmax-i)+1, column=17).value)

                    self.GE2depart=str(worksheet.cell(row=(rowmax-i)+1, column=20).value)

                    self.GE3depart=str(worksheet.cell(row=(rowmax-i)+1, column=23).value)

                    self.GE4depart=str(worksheet.cell(row=(rowmax-i)+1, column=26).value)

                    self.MP1arrive=str(worksheet.cell(row=(rowmax), column=5).value)

                    self.MP2arrive=str(worksheet.cell(row=(rowmax), column=8).value)

                    self.MP3arrive=str(worksheet.cell(row=(rowmax), column=11).value)

                    self.MP4arrive=str(worksheet.cell(row=(rowmax), column=14).value)
                    
                    self.GE1arrive=str(worksheet.cell(row=(rowmax), column=17).value)

                    self.GE2arrive=str(worksheet.cell(row=(rowmax), column=20).value)

                    self.GE3arrive=str(worksheet.cell(row=(rowmax), column=23).value)

                    self.GE4arrive=str(worksheet.cell(row=(rowmax), column=26).value)


                    self.MP1traverse=str(float(self.MP1arrive)-float(self.MP1depart))
                    
                    self.MP2traverse=str(float(self.MP2arrive)-float(self.MP2depart))

                    self.MP3traverse=str(float(self.MP3arrive)-float(self.MP3depart))

                    self.MP4traverse=str(float(self.MP4arrive)-float(self.MP4depart))
                    
                    self.GE1traverse=str(float(self.GE1arrive)-float(self.GE1depart))

                    self.GE2traverse=str(float(self.GE2arrive)-float(self.GE2depart))

                    self.GE3traverse=str(float(self.GE3arrive)-float(self.GE3depart))

                    self.GE4traverse=str(float(self.GE4arrive)-float(self.GE4depart))
                    print("traversé")
                    

                #mettre dans l'exel ancien valeur i

                elif str(worksheet.cell(row=rowmax-i, column=2).value)=="quai" and traversé_quai_found==1:
                    traversé_quai_found+=1
                    print(str(worksheet.cell(row=(rowmax-i)+1, column=5).value))
                    MP1quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=5).value)
                    MP2quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=8).value)
                    MP3quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=11).value)
                    MP4quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=14).value)
                    GE1quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=17).value)
                    GE2quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=20).value)
                    GE3quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=23).value)
                    GE4quai_debut_traverse=str(worksheet.cell(row=(rowmax-i)+1, column=26).value)


                    MP1quai_fin_quai=str(worksheet.cell(row=(rowmax-i)-1, column=5).value)
                    print("MP1quai_fin_quai",MP1quai_fin_quai)
                    MP2quai_fin_quai =str(worksheet.cell(row=(rowmax-i)-1, column=8).value)
                    MP3quai_fin_quai =str(worksheet.cell(row=(rowmax-i)-1, column=11).value)
                    MP4quai_fin_quai =str(worksheet.cell(row=(rowmax-i)-1, column=14).value)
                    GE1quai_fin_quai =str(worksheet.cell(row=(rowmax-i)-1, column=17).value)
                    GE2quai_fin_quai =str(worksheet.cell(row=(rowmax-i)-1, column=20).value)
                    GE3quai_fin_quai = str(worksheet.cell(row=(rowmax-i)-1, column=23).value)
                    GE4quai_fin_quai = str(worksheet.cell(row=(rowmax-i)-1, column=26).value)
                    try:
                        self.MP1quai=str(float(MP1quai_debut_traverse)-float(MP1quai_fin_quai))
                        print("float(MP1quai_debut_traverse)",float(MP1quai_debut_traverse))
                        self.MP2quai=str(float(MP2quai_debut_traverse)-float(MP2quai_fin_quai))

                        self.MP3quai=str(float(MP3quai_debut_traverse)-float(MP3quai_fin_quai))

                        self.MP4quai=str(float(MP4quai_debut_traverse)-float(MP4quai_fin_quai))
                        
                        self.GE1quai=str(float(GE1quai_debut_traverse)-float(GE1quai_fin_quai))

                        self.GE2quai=str(float(GE2quai_debut_traverse)-float(GE2quai_fin_quai))

                        self.GE3quai=str(float(GE3quai_debut_traverse)-float(GE3quai_fin_quai))

                        self.GE4quai=str(float(GE4quai_debut_traverse)-float(GE4quai_fin_quai))
                    except:
                        self.MP1quai=0

                        self.MP2quai=0

                        self.MP3quai=0

                        self.MP4quai=0

                        self.GE1quai=0

                        self.GE2quai=0

                        self.GE3quai=0

                        self.GE4quai=0


                if traversé_quai_found==2:
                    break
               
        refresh()


        workbook.save("fichier_excel/consommation.xlsx")
        
        def update():
            refresh()
            self.afficheur_consommation_MP1_Carburant1_depart1.configure(text= self.MP1depart)
            self.afficheur_consommation_MP1_Carburant1_arrive1.configure(text= self.MP1arrive)
            self.afficheur_consommation_MP1_traversé.configure(text= self.MP1traverse)
            self.afficheur_consommation_MP1_quai.configure(text= self.MP1quai)
            self.after(5000,update)
        def quaie_depart():
            
            workbook = load_workbook('fichier_excel/consommation.xlsx')
            worksheet = workbook.active
            worksheet.cell(row=1, column=2, value="traverse")
            workbook.save('fichier_excel/consommation.xlsx')
            workbook.close()
            refresh()
            self.afficheur_consommation_MP1_Carburant1_depart1.configure(text= self.MP1depart)
            self.afficheur_consommation_MP1_Carburant1_arrive1.configure(text= self.MP1arrive)
            self.afficheur_consommation_MP1_traversé.configure(text= self.MP1traverse)
            self.afficheur_consommation_MP1_quai.configure(text= self.MP1quai)
            self.after(5000,update)
        def arrivé_quai():
            workbook = load_workbook('fichier_excel/consommation.xlsx')
            worksheet = workbook.active
            worksheet.cell(row=1, column=2, value="quai")
            workbook.save('fichier_excel/consommation.xlsx')
            workbook.close()
            refresh()
            self.afficheur_consommation_MP1_Carburant1_depart1.configure(text= self.MP1depart)
            self.afficheur_consommation_MP1_Carburant1_arrive1.configure(text= self.MP1arrive)
            self.afficheur_consommation_MP1_traversé.configure(text= self.MP1traverse)
            self.afficheur_consommation_MP1_quai.configure(text= self.MP1quai)
            self.after(5000,update)

        def button_transfert_historique():
            import os
            

        def button_consommation():
            import os
            os.system('cd fichier_excel')
            os.system('libreoffice --calc [consommation.xlsx]')
            self.after(5000,update)
        
        self.Label_Titre=customtkinter.CTkLabel(master=self, width = 10, height = 150,text="Valeur de consommation",font=("Times", 30))
        
        self.Label_MP1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="MP1",font=("Times", 20))
        self.Label_MP2=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="MP2",font=("Times", 20))
        self.Label_MP3=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="MP3",font=("Times", 20))
        self.Label_MP4=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="MP4",font=("Times", 20))

        self.Label_GE1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="GE1",font=("Times", 20))
        self.Label_GE2=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="GE2",font=("Times", 20))
        self.Label_GE3=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="GE3",font=("Times", 20))
        self.Label_GE4=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="GE4",font=("Times", 20))


        #self.Label_carburant1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text=carburant1,font=("Times", 20))
        #self.Label_carburant1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="fuel",font=("Times", 20))
        self.Label_depart1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="Depart",font=("Times", 20))
        self.Label_arrive1=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="Arrive",font=("Times", 20))
        self.Label_consommation_traverse=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="Label consommation traverse",font=("Times", 20))
        self.Label_consommation_a_quai=customtkinter.CTkLabel(master=self, width = 10, height = 30,text="Label consommation a quai",font=("Times", 20))



        self.boutton_retour = customtkinter.CTkButton(master=self,width= 50,height= 20, text="retour",font=("Times", 20), command=retour_Window_menu)
        self.boutton_depart = customtkinter.CTkButton(master=self,width= 50,height= 20, text="boutton depart",font=("Times", 20), command=quaie_depart)
        self.boutton_arrivé = customtkinter.CTkButton(master=self,width= 50,height= 20, text="boutton arrivé",font=("Times", 20), command=arrivé_quai)
 
        self.widget_transfert_historique = customtkinter.CTkButton(master=self,width= 160,height= 34, text="transfert_historique",font=("Times", 20), command=button_transfert_historique)
        self.widget_historique = customtkinter.CTkButton(master=self,width= 160,height= 34, text="historique",font=("Times", 20), command=button_consommation)
        

      
        self.Label_Titre.grid(row=0, column=2, padx=5, pady=(5, 5),sticky="ew")
        
        #self.Label_Titre.grid(row=0, column=[1,2], padx=5, pady=(5, 5),sticky="ew")
        
        #self.Label_carburant1.grid(row=2, column=1, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_depart1.grid(row=3, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.Label_arrive1.grid(row=3, column=2, padx=5, pady=(5, 5),sticky="ew")

        self.Label_consommation_traverse.grid(row=3, column=3, padx=5, pady=(5, 5),sticky="ew")
        self.Label_consommation_a_quai.grid(row=3, column=4, padx=5, pady=(5, 5),sticky="ew")


        self.Label_consommation_traverse.grid(row=3, column=3, padx=5, pady=(5, 5),sticky="ew")
        self.Label_consommation_a_quai.grid(row=3, column=4, padx=5, pady=(5, 5),sticky="ew")


        #Carburant1
        self.afficheur_consommation_MP1_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP1depart,font=("Times", 20))
        self.afficheur_consommation_MP1_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP1arrive,font=("Times", 20))
        self.afficheur_consommation_MP2_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP2depart,font=("Times", 20))
        self.afficheur_consommation_MP2_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP2arrive,font=("Times", 20))

        self.afficheur_consommation_MP3_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP3depart,font=("Times", 20))
        self.afficheur_consommation_MP3_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP3arrive,font=("Times", 20))
        self.afficheur_consommation_MP4_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP4depart,font=("Times", 20))
        self.afficheur_consommation_MP4_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP4arrive,font=("Times", 20))


        self.afficheur_consommation_GE1_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE1depart,font=("Times", 20))
        self.afficheur_consommation_GE1_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE1arrive,font=("Times", 20))
        self.afficheur_consommation_GE2_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE2depart,font=("Times", 20))
        self.afficheur_consommation_GE2_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE2arrive,font=("Times", 20))

        self.afficheur_consommation_GE3_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE3depart,font=("Times", 20))
        self.afficheur_consommation_GE3_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE3arrive,font=("Times", 20))
        self.afficheur_consommation_GE4_Carburant1_depart1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE4depart,font=("Times", 20))
        self.afficheur_consommation_GE4_Carburant1_arrive1= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE4arrive,font=("Times", 20))
        
        #afficheur consomation de la traversé
        self.afficheur_consommation_MP1_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP1traverse,font=("Times", 20))
        self.afficheur_consommation_MP2_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP2traverse,font=("Times", 20))

        self.afficheur_consommation_MP3_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP3traverse,font=("Times", 20))
        self.afficheur_consommation_MP4_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP4traverse,font=("Times", 20))


        self.afficheur_consommation_GE1_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE1traverse,font=("Times", 20))
        self.afficheur_consommation_GE2_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE2traverse,font=("Times", 20))

        self.afficheur_consommation_GE3_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE3traverse,font=("Times", 20))
        self.afficheur_consommation_GE4_traversé= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE4traverse,font=("Times", 20))

        #consommation a quai
        self.afficheur_consommation_MP1_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP1quai,font=("Times", 20))
        self.afficheur_consommation_MP2_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP2quai,font=("Times", 20))

        self.afficheur_consommation_MP3_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP3quai,font=("Times", 20))
        self.afficheur_consommation_MP4_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.MP4quai,font=("Times", 20))


        self.afficheur_consommation_GE1_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE1quai,font=("Times", 20))
        self.afficheur_consommation_GE2_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE2quai,font=("Times", 20))

        self.afficheur_consommation_GE3_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE3quai,font=("Times", 20))
        self.afficheur_consommation_GE4_quai= customtkinter.CTkLabel(master=self, width = 10, height = 30,text=self.GE4quai,font=("Times", 20))


        #GRID consommation
        #A finir
        try:
            from configuration import window_configuration
        except:
            print("error importation configuration")
            
        window_configuration=window_configuration()
        liste_d_address=window_configuration.listconfig
        
        self.afficheur_consommation_MP1_Carburant1_depart1.grid(row=4, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP1_Carburant1_arrive1.grid(row=4, column=2, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP2_Carburant1_depart1.grid(row=5, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP2_Carburant1_arrive1.grid(row=5, column=2, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP3_Carburant1_depart1.grid(row=6, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP3_Carburant1_arrive1.grid(row=6, column=2, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP4_Carburant1_depart1.grid(row=7, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_MP4_Carburant1_arrive1.grid(row=7, column=2, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE1_Carburant1_depart1.grid(row=8, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE1_Carburant1_arrive1.grid(row=8, column=2, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE2_Carburant1_depart1.grid(row=9, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE2_Carburant1_arrive1.grid(row=9, column=2, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE3_Carburant1_depart1.grid(row=10, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE3_Carburant1_arrive1.grid(row=10, column=2, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE4_Carburant1_depart1.grid(row=11, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.afficheur_consommation_GE4_Carburant1_arrive1.grid(row=11, column=2, padx=5, pady=(5, 5),sticky="ew")
    
        #GRID consommation de la traversé
        self.afficheur_consommation_MP1_traversé.grid(row=4, column=3, padx=5, pady=(5, 5),sticky="ew")
        
        self.afficheur_consommation_MP2_traversé.grid(row=5, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP3_traversé.grid(row=6, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP4_traversé.grid(row=7, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE1_traversé.grid(row=8, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE2_traversé.grid(row=9, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE3_traversé.grid(row=10, column=3, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE4_traversé.grid(row=11, column=3, padx=5, pady=(5, 5),sticky="ew")

        #GRID QUAI
        self.afficheur_consommation_MP1_quai.grid(row=4, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP2_quai.grid(row=5, column=4, padx=5, pady=(5, 5),sticky="ew")

        
        self.afficheur_consommation_MP3_quai.grid(row=6, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP4_quai.grid(row=7, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE1_quai.grid(row=8, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE2_quai.grid(row=9, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE3_quai.grid(row=10, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE4_quai.grid(row=11, column=4, padx=5, pady=(5, 5),sticky="ew")


        #GRID label comsomation

        self.Label_MP1.grid(row=4, column=0, padx=5, pady=(5, 5),sticky="ew")
        self.Label_MP2.grid(row=5, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_MP3.grid(row=6, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_MP4.grid(row=7, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_GE1.grid(row=8, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_GE2.grid(row=9, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_GE3.grid(row=10, column=0, padx=5, pady=(5, 5),sticky="ew")
        
        self.Label_GE4.grid(row=11, column=0, padx=5, pady=(5, 5),sticky="ew")
        






        self.afficheur_consommation_MP1_quai.grid(row=4, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP2_quai.grid(row=5, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP3_quai.grid(row=6, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_MP4_quai.grid(row=7, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE1_quai.grid(row=8, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE2_quai.grid(row=9, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE3_quai.grid(row=10, column=4, padx=5, pady=(5, 5),sticky="ew")

        self.afficheur_consommation_GE4_quai.grid(row=11, column=4, padx=5, pady=(5, 5),sticky="ew")




        self.boutton_retour.grid(row=12, column=0, padx=5, pady=(5, 5),sticky="ew")
        self.boutton_depart.grid(row=12, column=1, padx=5, pady=(5, 5),sticky="ew")
        self.boutton_arrivé.grid(row=12, column=2, padx=5, pady=(5, 5),sticky="ew")
        self.widget_historique.grid(row=12, column=3, padx=5, pady=(5, 5),sticky="ew")
        self.widget_transfert_historique.grid(row=12, column=4, padx=5, pady=(5, 5),sticky="ew")


        self.after(5000,update)
        self.mainloop()

#window_Valeur_de_consommation("Admin")


