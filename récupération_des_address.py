import minimalmodbus # Don't forget to import the library!!
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import subprocess
from datetime import datetime
import time

def getalladdress(scan=False):
    liste_address=[]
    i=0
    workbook = load_workbook('fichier_excel/recuperation_Address.xlsx')
    worksheet = workbook.active
    liste_address=[]
    #or str(worksheet.cell(row=2, column=1).value)=="" or scan==False
    #if str(worksheet.cell(row=2, column=1).value) is not None or str(worksheet.cell(row=2, column=1).value)!="None" :
    if worksheet.cell(row=2, column=1).value is not None or str(worksheet.cell(row=2, column=1).value)!="None" :
        rowmax=worksheet.max_row
        for i in range(1,rowmax):
            liste_address.append(str(worksheet.cell(row=1+i, column=1).value))
        
        workbook.save("fichier_excel/recuperation_Address.xlsx")
        workbook.close()
        return(liste_address)
    else:
        workbook.save("fichier_excel/recuperation_Address.xlsx")
        workbook.close()
        workbook = load_workbook('fichier_excel/Etat Communication.xlsx')
        worksheet = workbook.active
        while str(worksheet.cell(row=1, column=2).value)=="Actif":
            time.sleep(0.5)
        worksheet.cell(row=1, column=2, value="Actif")

        for addr in range(1,20):
            mb_address = addr # Modbus address of sensor
            try:
                
                rs485communication = minimalmodbus.Instrument('/dev/ttyUSB0',mb_address)	# Make an "instrument" object called rs485communication (port name, slave address (in decimal))
                rs485communication.serial.baudrate = 19200				# BaudRate
                rs485communication.serial.bytesize = 8					# Number of data bits to be requested
                rs485communication.serial.parity = minimalmodbus.serial.PARITY_EVEN	# Parity Setting here is NONE but can be ODD or EVEN
                rs485communication.serial.stopbits = 1					# Number of stop bits
                rs485communication.serial.timeout  = 0.5					# Timeout time in seconds
                rs485communication.mode = minimalmodbus.MODE_RTU		# Mode to be used (RTU or ascii mode)
                
                #rs485communication.clear_buffers_before_each_transaction = True
                #rs485communication.close_port_after_each_call = True
                
                
                if rs485communication.read_float(2006,3,2,3) is not None or rs485communication.read_float(2006,3,2,3)==0:
                    liste_address.append(addr)

                workbook = load_workbook('fichier_excel/recuperation_Address.xlsx')
                worksheet = workbook.active
                
                worksheet.delete_cols(1)
                for i in range(len(liste_address)):
                    worksheet.delete_cols(1)
                    worksheet.append({'A': "address:"})
                    worksheet.append({'A': liste_address[i]})
                    workbook.save("fichier_excel/recuperation_Address.xlsx")

            except IOError:
                print("Failed to read the address",addr," from instrument")

        worksheet.cell(row=1, column=2, value="Inactif")


        return(liste_address)

#getalladdress=getalladdress()
            
