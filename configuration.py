import minimalmodbus # Don't forget to import the library!!
import customtkinter
import os
from PIL import Image
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import subprocess
from datetime import datetime
import os
import openpyxl
import filelock
import time






file_path_recuperation_Address = 'fichier_excel/recuperation_Address.xlsx'
sheet_name = "Feuille1"


list_time=[datetime.now()]

def lock_excel_file():
    # Create a file lock
    file_lock = filelock.FileLock(file_path_recuperation_Address + ".lock")
    max_attempts = 1  # Maximum attempts to acquire the lock
    attempt = 0
    
    while attempt < max_attempts:
        try:
            # Acquire the lock with a timeout of 20 seconds
            with file_lock.acquire(timeout=20):
                # Perform operations on the Excel file
                wb = openpyxl.load_workbook(file_path_recuperation_Address)
                sheet = wb.active
                from récupération_des_address import getalladdress
                getalladdress=getalladdress()
                getalladdress = [str(x) for x in getalladdress]
                return(getalladdress)
                
        except filelock.Timeout:
            attempt += 1
            if attempt < max_attempts:
                time.sleep(1)  # Optional: Add a delay before retrying

        if attempt == max_attempts:
            print("Maximum attempts reached. Lock could not be acquired.")


class ScrollableLabelButtonFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master,utilisateur, command=None, **kwargs):
        super().__init__(master, **kwargs)
        self.utilisateur=utilisateur
        self.grid_columnconfigure((0,1,2), weight=1)
        self.command = command
        self.master= master
        self.radiobutton_variable = customtkinter.StringVar()
        self.label_list = []
        self.button_list = []
        self.nameEntry_list= []
    def retour_Window_menu(self):
        from Menu import window_Menu
        
        self.master.destroy()
        window_Menuu=window_Menu(self.utilisateur)
        #window_Menuu.mainloop()

    def scan_capteur(self,liste_adress, item):

        getalladdress=getalladdress(True)
        getalladdress = [str(x) for x in getalladdress]
        for label,nameEntry_list, button in zip(self.label_list,self.nameEntry_list, self.button_list):
            if item == label.cget("text"):
                
                nameEntry_list.set(liste_adress)
                return

    def add_item(self, item,input,image=None):
        def writte_time(datetime):
            process = subprocess.Popen(['sudo', 'date','-s' ,datetime],
                     stdout=subprocess.PIPE,
                     stderr=subprocess.PIPE)
            subprocess.run()
            #subprocess.run(["/path/to/your/shell/script","arguments"], shell=True)
        def command_add_item(label):
            if label.cget("text")=="\t\t\t\t\t\t\t Date time format :YYYY-MM-DD HH:MM:SS ": 
                writte_time(nameEntry.cget("text"))
            if label.cget("text")=="\t\t\t\t\t\t\t retour":
                self.retour_Window_menu()
            if label.cget("text")=="\t\t\t\t\t\t\t Scan des capteurs":
                self.scan_capteur()
        label = customtkinter.CTkLabel(self, text=item, image=image, compound="left", padx=5, anchor="w")
        button = customtkinter.CTkButton(self, text="Confirmé", width=100, height=24,command=lambda: command_add_item(label))
        nameEntry = customtkinter.CTkEntry(master=self,placeholder_text=input,font=("Times", 15))

        label.grid(row=len(self.label_list), column=0, padx=5,pady=(0, 10), sticky="nsew")
        if input!=" ":
            nameEntry.grid(row=len(self.nameEntry_list), column=1,padx=5, pady=(0, 10),sticky="nsew")
        button.grid(row=len(self.button_list), column=2,padx=5, pady=(0, 10),sticky="nsew")

        self.label_list.append(label)
        self.button_list.append(button)
        self.nameEntry_list.append(nameEntry)
        
    def add_itemtitle(self, item, image=None):
        label = customtkinter.CTkLabel(self, text=item, image=image, compound="left", padx=5, anchor="w")
        labelremplissage = customtkinter.CTkLabel(self, text=None, image=image, compound="left", padx=5, anchor="w")
        label.grid(row=len(self.label_list), column=0,padx=2, pady=(0, 10), sticky="w")
        labelremplissage.grid(row=len(self.label_list), column=1,padx=5, pady=(0, 10), sticky="w")
        labelremplissage.grid(row=len(self.label_list), column=2,padx=5, pady=(0, 10), sticky="w")
        self.label_list.append(label)
        self.nameEntry_list.append(labelremplissage)
        self.button_list.append(labelremplissage)
        

    def modify_item(self,new_name_entry, item):
        for label,nameEntry_list, button in zip(self.label_list,self.nameEntry_list, self.button_list):
            if item == label.cget("text"):
                nameEntry_list.set(new_name_entry)
                return
            


    def modify_all(self,MP1IN,MP1OUT,MP2IN,MP2OUT,MP3IN,MP3OUT,MP4IN,MP4OUT,GE1IN,GE1OUT,GE2IN,GE2OUT,GE3IN,GE3OUT,GE4IN,GE4OUT):
        
        self.modify_item(MP1IN,"\t\t\t\t\t\t\t MP1IN")
        self.modify_item(MP1IN,"\t\t\t\t\t\t\t MP1OUT")
        self.modify_item(MP2IN,"\t\t\t\t\t\t\t MP2IN")
        self.modify_item(MP2OUT,"\t\t\t\t\t\t\t MP2OUT")
        self.modify_item(MP3IN,"\t\t\t\t\t\t\t MP3IN")
        self.modify_item(MP3IN,"\t\t\t\t\t\t\t MP3OUT")
        self.modify_item(MP4IN,"\t\t\t\t\t\t\t MP4IN")
        self.modify_item(MP4IN,"\t\t\t\t\t\t\t MP4OUT")

        self.modify_item(GE1IN,"\t\t\t\t\t\t\t GE1IN")
        self.modify_item(GE1IN,"\t\t\t\t\t\t\t GE1OUT")
        self.modify_item(GE2IN,"\t\t\t\t\t\t\t GE2IN")
        self.modify_item(GE2IN,"\t\t\t\t\t\t\t GE2OUT")
        self.modify_item(GE3IN,"\t\t\t\t\t\t\t GE3IN")
        self.modify_item(GE3IN,"\t\t\t\t\t\t\t GE3OUT")
        self.modify_item(GE4IN,"\t\t\t\t\t\t\t GE4IN")
        self.modify_item(GE4IN,"\t\t\t\t\t\t\t GE4OUT")


    def add_liste_a_choix(self,Label,Address,liste_des_choix):

        def write_address():

            file_path_Configuration_capteur = 'fichier_excel/Configuration_capteur.xlsx'
            sheet_name = "Feuille1"
            # Create a file lock
            file_lock = filelock.FileLock(file_path_Configuration_capteur + ".lock")
            max_attempts = 1  # Maximum attempts to acquire the lock
            attempt = 0
                
            while attempt < max_attempts:
                try:
                    # Acquire the lock with a timeout of 20 seconds
                    with file_lock.acquire(timeout=20):
                        # Perform operations on the Excel file
                        
                        workbook = load_workbook(file_path_Configuration_capteur)
                        worksheet = workbook.active

                        address_capteur=liste_a_choix.get()
                        resultat=liste_a_choix.get()
                        if resultat is not None:
                            
                            if Label.find("self.bateau")!=-1:
                                self.Bateau=resultat
                                
                                if self.Bateau=="Armorique":
                                    colonne=5
                                elif self.Bateau=="Pont-Aven":
                                    colonne=6
                                elif self.Bateau=="Mont ST Michel":
                                    colonne=7
                                elif self.Bateau=="Barfleur":
                                    colonne=8
                                elif self.Bateau=="Bretagne":
                                    colonne=9
                                elif self.Bateau=="Cotentin":
                                    colonne=10
                                elif self.Bateau=="Galicia":
                                    colonne=11
                                elif self.Bateau=="Normandie":
                                    colonne=12
                                elif self.Bateau=="Salamanca":
                                    colonne=13
                                elif self.Bateau=="Pelican":
                                    colonne=14
                                worksheet.cell(row=1, column=2, value=str(worksheet.cell(row=1, column=colonne).value))
                                worksheet.cell(row=2, column=2, value=str(worksheet.cell(row=2, column=colonne).value))
                                worksheet.cell(row=3, column=2, value=str(worksheet.cell(row=3, column=colonne).value))
                                worksheet.cell(row=4, column=2, value=str(worksheet.cell(row=4, column=colonne).value))
                                worksheet.cell(row=5, column=2, value=str(worksheet.cell(row=5, column=colonne).value))
                                worksheet.cell(row=6, column=2, value=str(worksheet.cell(row=6, column=colonne).value))
                                worksheet.cell(row=7, column=2, value=str(worksheet.cell(row=7, column=colonne).value))
                                worksheet.cell(row=8, column=2, value=str(worksheet.cell(row=8, column=colonne).value))
                                worksheet.cell(row=9, column=2, value=str(worksheet.cell(row=9, column=colonne).value))
                                worksheet.cell(row=10, column=2, value=str(worksheet.cell(row=10, column=colonne).value))
                                worksheet.cell(row=11, column=2, value=str(worksheet.cell(row=11, column=colonne).value))
                                worksheet.cell(row=12, column=2, value=str(worksheet.cell(row=11, column=colonne).value))
                                worksheet.cell(row=13, column=2, value=str(worksheet.cell(row=13, column=colonne).value))
                                worksheet.cell(row=14, column=2, value=str(worksheet.cell(row=14, column=colonne).value))
                                worksheet.cell(row=15, column=2, value=str(worksheet.cell(row=15, column=colonne).value))
                                worksheet.cell(row=16, column=2, value=str(worksheet.cell(row=16, column=colonne).value))
                                worksheet.cell(row=17, column=2, value=str(worksheet.cell(row=17, column=colonne).value))
                        self.Bateau=str(worksheet.cell(row=1, column=2).value)
                        if self.Bateau=="Armorique":
                            colonne=5
                        elif self.Bateau=="Pont-Aven":
                            colonne=6
                        elif self.Bateau=="Mont ST Michel":
                            colonne=7
                        elif self.Bateau=="Barfleur":
                            colonne=8
                        elif self.Bateau=="Bretagne":
                            colonne=9
                        elif self.Bateau=="Cotentin":
                            colonne=10
                        elif self.Bateau=="Galicia":
                            colonne=11
                        elif self.Bateau=="Normandie":
                            colonne=12
                        elif self.Bateau=="Salamanca":
                            colonne=13
                        elif self.Bateau=="Pelican":
                            colonne=14
                        if Label.find("MP1IN")!=-1:
                            worksheet.cell(row=2, column=2, value=address_capteur)
                            self.MP1IN=worksheet.cell(row=2, column=colonne, value=address_capteur)
                        elif Label.find("MP1OUT")!=-1:
                            worksheet.cell(row=3, column=2, value=address_capteur)
                            self.MP1OUT=worksheet.cell(row=3, column=colonne, value=address_capteur)
                        elif Label.find("MP2IN")!=-1:
                            worksheet.cell(row=4, column=2, value=address_capteur)
                            self.MP2IN=worksheet.cell(row=4, column=colonne, value=address_capteur)
                        elif Label.find("MP2OUT")!=-1:
                            worksheet.cell(row=5, column=2, value=address_capteur)
                            self.MP2OUT=worksheet.cell(row=5, column=colonne, value=address_capteur)
                        elif Label.find("MP3IN")!=-1:
                            worksheet.cell(row=6, column=colonne, value=address_capteur)
                            self.MP3IN=worksheet.cell(row=6, column=2, value=address_capteur)
                        elif Label.find("MP3OUT")!=-1:
                            worksheet.cell(row=7, column=2, value=address_capteur)
                            self.MP3OUT=worksheet.cell(row=7, column=colonne, value=address_capteur)
                        elif Label.find("MP4IN")!=-1:
                            worksheet.cell(row=8, column=2, value=address_capteur)
                            self.MP4IN=worksheet.cell(row=8, column=colonne, value=address_capteur)
                        elif Label.find("MP4OUT")!=-1:
                            worksheet.cell(row=9, column=2, value=address_capteur)
                            self.MP4OUT=worksheet.cell(row=9, column=colonne, value=address_capteur)
                        elif Label.find("GE1IN")!=-1:
                            worksheet.cell(row=10, column=2, value=address_capteur)
                            self.GE1IN=worksheet.cell(row=10, column=colonne, value=address_capteur)
                        elif Label.find("GE1OUT")!=-1:
                            worksheet.cell(row=11, column=2, value=address_capteur)
                            self.GE1OUT=worksheet.cell(row=11, column=colonne, value=address_capteur)
                        elif Label.find("GE2IN")!=-1:
                            worksheet.cell(row=12, column=2, value=address_capteur)
                            self.GE2IN=worksheet.cell(row=12, column=colonne, value=address_capteur)
                        elif Label.find("GE2OUT")!=-1:
                            worksheet.cell(row=13, column=2, value=address_capteur)
                            self.GE2OUT=worksheet.cell(row=13, column=colonne, value=address_capteur)
                        elif Label.find("GE2OUT")!=-1:
                            worksheet.cell(row=14, column=2, value=address_capteur)
                            self.GE3IN=worksheet.cell(row=14, column=colonne, value=address_capteur)
                        elif Label.find("GE3OUT")!=-1:
                            worksheet.cell(row=15, column=2, value=address_capteur)
                            self.GE3OUT=worksheet.cell(row=15, column=colonne, value=address_capteur)
                        elif Label.find("GE4IN")!=-1:
                            worksheet.cell(row=16, column=2, value=address_capteur)
                            self.GE4IN=worksheet.cell(row=16, column=colonne, value=address_capteur)
                        elif Label.find("GE4OUT")!=-1:
                            worksheet.cell(row=17, column=2, value=address_capteur)
                            self.GE4OUT=worksheet.cell(row=17, column=colonne, value=address_capteur)
                        elif Label.find("retour")!=-1:
                            worksheet.cell(row=17, column=2, value=address_capteur)
                            self.GE4OUT=worksheet.cell(row=17, column=colonne, value=address_capteur)

                        
                        self.MP1IN=str(worksheet.cell(row=2, column=2).value)

                        self.MP1OUT=str(worksheet.cell(row=3, column=2).value)
                        
                        self.MP2IN=str(worksheet.cell(row=4, column=2).value)

                        self.MP2OUT=str(worksheet.cell(row=5, column=2).value)
                        
                        self.MP3IN=str(worksheet.cell(row=6, column=2).value)

                        self.MP3OUT=str(worksheet.cell(row=7, column=2).value)

                        self.MP4IN=str(worksheet.cell(row=8, column=2).value)

                        self.MP4OUT=str(worksheet.cell(row=9, column=2).value)
                        
                        self.GE1IN=str(worksheet.cell(row=10, column=2).value)

                        self.GE1OUT=str(worksheet.cell(row=11, column=2).value)

                        self.GE2IN=str(worksheet.cell(row=12, column=2).value)

                        self.GE2OUT=str(worksheet.cell(row=13, column=2).value)
                        
                        self.GE3IN=str(worksheet.cell(row=14, column=2).value)

                        self.GE3OUT=str(worksheet.cell(row=15, column=2).value)

                        self.GE4IN=str(worksheet.cell(row=16, column=2).value)

                        self.GE4OUT=str(worksheet.cell(row=17, column=2).value)    
                        attempt += 1
                        workbook.save(file_path_Configuration_capteur)
                except filelock.Timeout:
                    attempt += 1
                    print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
                    if attempt < max_attempts:
                        time.sleep(1)  # Optional: Add a delay before retrying

                if attempt == max_attempts:
                    print("Maximum attempts reached.")
                    

            self.modify_all(self.MP1IN,self.MP1OUT,self.MP2IN,self.MP2OUT,self.MP3IN,self.MP3OUT,self.MP4IN,self.MP4OUT,self.GE1IN,self.GE1OUT,self.GE2IN,self.GE2OUT,self.GE3IN,self.GE3OUT,self.GE4IN,self.GE4OUT)
        label=customtkinter.CTkLabel(self, text= Label ,compound="left", padx=5, anchor="w")
        liste_a_choix = customtkinter.CTkComboBox(self,values=liste_des_choix)
        button = customtkinter.CTkButton(self, text="Confirmé",command=write_address, width=100, height=24)

        label.grid(row=len(self.label_list), column=0,padx=5, pady=(0, 10), sticky="w")
        liste_a_choix.grid(row=len(self.label_list), column=1, padx=5, pady=(0, 10), sticky="nsew")
        button.grid(row=len(self.label_list), column=2,padx=5, pady=(0, 10), sticky="nsew")

        self.label_list.append(label)
        self.nameEntry_list.append(liste_a_choix)
        self.button_list.append(button)
        liste_a_choix.set(Address)

class window_configuration(customtkinter.CTk):
    def __init__(self,utilisateur):
        super().__init__()

        
        self.utilisateur=utilisateur
        file_path_Configuration_capteur = 'fichier_excel/Configuration_capteur.xlsx'
        workbook = load_workbook(file_path_Configuration_capteur)
        worksheet = workbook.active
        sheet_name = "Feuille1"
        # Create a file lock
        file_lock = filelock.FileLock(file_path_Configuration_capteur + ".lock")
        print("here def lock_excel")
        max_attempts = 1  # Maximum attempts to acquire the lock
        attempt = 0
        
        while attempt < max_attempts:
            try:
                print("here while attempt first")
                # Acquire the lock with a timeout of 20 seconds
                with file_lock.acquire(timeout=20):
                    print(f"Lock acquired for first {file_path_Configuration_capteur}")
                    
                    # Perform operations on the Excel file
                    wb = openpyxl.load_workbook(file_path_Configuration_capteur)
                    sheet = wb.active


                    self.title("Configuration des capteurs")
                    self.columnconfigure(0, weight=1)

                    
                    self.Bateau=str(worksheet.cell(row=1, column=2).value)
                    colonne=5
                    if self.Bateau=="Armorique":
                        colonne=5
                    elif self.Bateau=="Pont-Aven":
                        colonne=6
                    elif self.Bateau=="Mont ST Michel":
                        colonne=7
                    elif self.Bateau=="Barfleur":
                        colonne=8
                    elif self.Bateau=="Bretagne":
                        colonne=9
                    elif self.Bateau=="Cotentin":
                        colonne=10
                    elif self.Bateau=="Galicia":
                        colonne=11
                    elif self.Bateau=="Normandie":
                        colonne=12
                    elif self.Bateau=="Salamanca":
                        colonne=13
                    elif self.Bateau=="Pelican":
                        colonne=14

                    def modify():
                        self.scrollable_radiobutton_frame.configure(label_text=200) 
                    def recuperationconfig():
                        self.MP1IN=str(worksheet.cell(row=2, column=2).value)

                        self.MP1OUT=str(worksheet.cell(row=3, column=2).value)
                        
                        self.MP2IN=str(worksheet.cell(row=4, column=2).value)

                        self.MP2OUT=str(worksheet.cell(row=5, column=2).value)
                        
                        self.MP3IN=str(worksheet.cell(row=6, column=2).value)

                        self.MP3OUT=str(worksheet.cell(row=7, column=2).value)

                        self.MP4IN=str(worksheet.cell(row=8, column=2).value)

                        self.MP4OUT=str(worksheet.cell(row=9, column=2).value)
                        
                        self.GE1IN=str(worksheet.cell(row=10, column=2).value)

                        self.GE1OUT=str(worksheet.cell(row=11, column=2).value)

                        self.GE2IN=str(worksheet.cell(row=12, column=2).value)

                        self.GE2OUT=str(worksheet.cell(row=13, column=2).value)
                        
                        self.GE3IN=str(worksheet.cell(row=14, column=2).value)

                        self.GE3OUT=str(worksheet.cell(row=15, column=2).value)

                        self.GE4IN=str(worksheet.cell(row=16, column=2).value)

                        self.GE4OUT=str(worksheet.cell(row=17, column=2).value)

                        self.listconfig={"MP1IN":self.MP1IN,"MP1OUT":self.MP1OUT,"MP2IN":self.MP2IN,"MP2OUT":self.MP2OUT,"MP3IN":self.MP3IN,"MP3OUT":self.MP3OUT,"MP4IN":self.MP4IN,"MP4OUT":self.MP4OUT,"GE1IN":self.GE1IN,"GE1OUT":self.GE1OUT,"GE2IN":self.GE2IN,"GE2OUT":self.GE2OUT,"GE3IN":self.GE3IN,"GE3OUT":self.GE3OUT,"GE4IN":self.GE4IN,"GE4OUT":self.GE4OUT}
                    recuperationconfig()
                    attempt += 1
                    workbook.save(file_path_Configuration_capteur)
                
            except filelock.Timeout:
                attempt += 1
                print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
                if attempt < max_attempts:
                    time.sleep(1)  # Optional: Add a delay before retrying

            if attempt == max_attempts:
                print("Maximum attempts reached.")



    def show(self):
        getalladdress=lock_excel_file()
        file_path_Configuration_capteur = 'fichier_excel/Configuration_capteur.xlsx'
        sheet_name = "Feuille1"
        # Create a file lock
        file_lock = filelock.FileLock(file_path_Configuration_capteur + ".lock")
        print("here def lock_excel")
        max_attempts = 1  # Maximum attempts to acquire the lock
        attempt = 0
        self.scrollable_label_button_frame = ScrollableLabelButtonFrame(master=self,utilisateur=self.utilisateur, width=900,height=800,command=self.radiobutton_frame_event, corner_radius=0)
        self.scrollable_label_button_frame.grid(row=0, column=0, padx=5, pady=0, sticky="nsew")
        self.scrollable_label_button_frame.add_itemtitle(f" Configuration des capteurs")
        self.scrollable_label_button_frame.add_itemtitle(f"\t Attribuer les address des capteurs a leurs roles respectif")
        listbateau=["Armorique","Pont-Aven","Mont ST Michel","Barfleur","Bretagne","Cotentin","Galicia","Normandie","Salamanca","Pelican"]
                    
        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t retour"," ") 
        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t bateau",self.Bateau,listbateau)
        self.scrollable_label_button_frame.configure()
        while attempt < max_attempts:
            try:

                # Acquire the lock with a timeout of 20 seconds
                with file_lock.acquire(timeout=20):

                    # Perform operations on the Excel file

                    getalladdress = [str(x) for x in getalladdress]


            #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Scan des capteurs"," ")

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP1IN",self.MP1IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP1OUT",self.MP1OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP2IN",self.MP2IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP2OUT",self.MP2OUT,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP3IN",self.MP3IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP3OUT",self.MP3OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP4IN",self.MP4IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t MP4OUT",self.MP4OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE1IN",self.GE1IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE1OUT",self.GE1OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE2IN",self.GE2IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE2OUT",self.GE2OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE3IN",self.GE3IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE3OUT",self.GE3OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE4IN",self.GE4IN,getalladdress)

                    self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t GE4OUT",self.GE4OUT,getalladdress)
                        
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Date time format :YYYY-MM-DD HH:MM:SS ",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t "," ")
                    attempt += 1        
            except filelock.Timeout:
                attempt += 1
                print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
            if attempt < max_attempts:
                time.sleep(1)  # Optional: Add a delay before retrying
                print("sleep")
            print("Excel file updated.")
            self.mainloop()


    def radiobutton_frame_event(self):
        print(f"radiobutton frame modified: {self.scrollable_radiobutton_frame.get_checked_item()}")

#window_configurationn= window_configuration("Admin")
#window_configurationn.show()

