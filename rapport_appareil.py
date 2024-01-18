#rapport_d_appareil
#
#import minimalmodbus
import customtkinter
import os
from PIL import Image
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import time
import openpyxl
import filelock

    
try:
    from windowerreur import window_erreur
except:
    print("error importation testwindowerreur")


class ScrollableLabelButtonFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master,communication,utilisateur, command=None, **kwargs):
        super().__init__(master, **kwargs)
        self.grid_columnconfigure((0,1,2), weight=1)
        self.utilisateur=utilisateur
        #self.grid_columnconfigure((0), weight=2)
        self.command = command
        self.communication=communication
        self.radiobutton_variable = customtkinter.StringVar()
        self.label_list = []
        self.button_list = []
        self.nameEntry_list= []

    def what_to_do():
        print("help")
    
    def add_item(self, item,input,registre=None,type=0,image=None):
        

        def refresh(text):

            file_path_Etat_Communication = 'fichier_excel/Etat Communication.xlsx'
            sheet_name = "Feuille1"

            
            def lock_excel_file():
                # Create a file lock
                file_lock = filelock.FileLock(file_path_Etat_Communication + ".lock")
                print("here def lock_excel")
                max_attempts = 1  # Maximum attempts to acquire the lock
                attempt = 0
                if text=="\t\t retour ":
                    print("retour")
                    from menu_rapport_d_appareil import window_Menu_rapport
                    self.destroy()
                    self.master.destroy()
                    window_Menu_rapportu=window_Menu_rapport(self.utilisateur)
                    window_Menu_rapportu.mainloop()
                while attempt < max_attempts:
                    try:
                        print("here while attempt first")
                        # Acquire the lock with a timeout of 20 seconds
                        with file_lock.acquire(timeout=20):
                            print(f"Lock acquired for first {file_path_Etat_Communication}")
                            
                            # Perform operations on the Excel file
                            wb = openpyxl.load_workbook(file_path_Etat_Communication)
                            sheet = wb.active
                            value=input.get()
                            if value is not None:
                                nameEntry.configure(placeholder_text= value)
                                nameEntry.delete(0,"end")
                                if value is not None and type==0:
                                    self.communication.write_register(registre,value)


                                elif value is not None and type==1:
                                    try:
                                        
                                        print("i am floatwrite")
                                        self.communication.write_float_sensor(registre,value)

                                        
                                    except:
                                        window_erreur=window_erreur()
                                        window_erreur
                                        window_erreur.mainloop()
                                        
                                elif value is not None and type==2:
                                    print("i am stringwrite")

                                    try:
                                        
                                        self.communication.write_string(registre,value)
                                        
                                    except:
                                        window_erreur=window_erreur()
                                        window_erreur
                                        window_erreur.mainloop()
                                attempt += 1
                    except filelock.Timeout:
                        attempt += 1
                        print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
                        if attempt < max_attempts:
                            time.sleep(1)  # Optional: Add a delay before retrying

                    if attempt == max_attempts:
                        print("Maximum attempts reached. Lock could not be acquired.")
            lock_excel_file()
        label = customtkinter.CTkLabel(self, text=item, image=image, compound="lef", padx=5, anchor="w")
        button = customtkinter.CTkButton(self, text="Confirmé",command=lambda: refresh(label.cget("text")),width=100, height=24,)
        nameEntry = customtkinter.CTkEntry(master=self,placeholder_text=input,font=("Times", 15))
        
        label.grid(row=len(self.label_list), column=0, padx=5,pady=(0, 10), sticky="nsew")
        nameEntry.grid(row=len(self.nameEntry_list), column=1,padx=5, pady=(0, 10),sticky="nsew")
        button.grid(row=len(self.button_list), column=2,padx=5, pady=(0, 10),sticky="nsew")

        self.label_list.append(label)
        self.button_list.append(button)
        self.nameEntry_list.append(nameEntry)
        
    def add_itemtitle(self, item, image=None):
        label = customtkinter.CTkLabel(self, text=item, image=image, compound="left", padx=5, anchor="w")
        labelremplissage = customtkinter.CTkLabel(self, text=None, image=image, compound="left", padx=5, anchor="w")
        label.grid(row=len(self.label_list), column=0,padx=2, pady=(0, 10), sticky="w")
        labelremplissage.grid(row=len(self.label_list), column=1,padx=5, pady=(0, 10), sticky="w")
        labelremplissage.grid(row=len(self.label_list), column=2,padx=5, pady=(0, 10), sticky="w")
        self.label_list.append(label)
        self.nameEntry_list.append(labelremplissage)
        self.button_list.append(labelremplissage)

    def add_liste_a_choix(self,Label,valeur_du_capteur,liste_des_choix,registre):



        def write_new_value(registre,rawvalue_a_envoyer):
            file_path_Etat_Communication = 'fichier_excel/Etat Communication.xlsx'
            file_path_consommation = 'fichier_excel/consommation.xlsx'
            sheet_name = "Feuille1"
            # Create a file lock
            file_lock = filelock.FileLock(file_path_Etat_Communication + ".lock")
            print("here def lock_excel")
            max_attempts = 1  # Maximum attempts to acquire the lock
            attempt = 0
            
            while attempt < max_attempts:
                try:
                    print("here while attempt first")
                    # Acquire the lock with a timeout of 20 seconds
                    with file_lock.acquire(timeout=20):
                        print(f"Lock acquired for first {file_path_Etat_Communication}")
                        
                        # Perform operations on the Excel file
                        wb = openpyxl.load_workbook(file_path_Etat_Communication)
                        sheet = wb.active


                    print("i am registerwrite")
                    def traduction_ordi_to_capt_Affecter_Numéro_de_diagnostic(Affecter_Numéro_de_diagnostic):
                        if Affecter_Numéro_de_diagnostic =="off":
                            return(0)
                        elif Affecter_Numéro_de_diagnostic =="Logbook entry only":
                            return(1)
                        elif Affecter_Numéro_de_diagnostic =="Warning":
                            return(2)
                        elif Affecter_Numéro_de_diagnostic =="Alarm":
                            return(3)
                    def traduction_ordi_to_capt_Device_reset(Device_reset):
                        if Device_reset =="Cancel":
                            return(0)
                        elif Device_reset =="Restart device":
                            return(1)
                        elif Device_reset =="To delivery settings":
                            return(2)
                    def traduction_ordi_to_capt_Aperçu_des_options_logiciels(Aperçu_des_options_logiciels):
                        if Aperçu_des_options_logiciels =="Heartbeat Monitoring":
                            return(16384)
                        elif Aperçu_des_options_logiciels =="Concentration":
                            return(4)
                        elif Aperçu_des_options_logiciels =="Heartbeat Verification":
                            return(32768)
                        elif Aperçu_des_options_logiciels =="Viscosity":
                            return(64)
                    def traduction_ordi_to_capt_Sauvegarde_permanente(Sauvegarde_permanente):
                        if Sauvegarde_permanente =="Off":
                            return(0)
                        elif Sauvegarde_permanente =="On":
                            return(1)
                        
                    def traduction_ordi_to_capt_Unité_de_débit_massique(Unité_de_débit_massique):
                        list=['g/s','g/min','g/h','g/d','kg/s','kg/min','kg/h','kg/d','t/s','t/min','t/h','t/d','oz/s','oz/min','oz/h','oz/d','lb/s','lb/min','lb/h','lb/d','STon/s','STon/min','STon/h','STon/d','User mass/s','User mass/min','User mass/h']  
                        return(list.index(Unité_de_débit_massique))
                    def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                        list=['g','kg','t','oz','lb','STon','User mass']
                        return(list.index(Unité_de_masse))
                    def traduction_ordi_to_capt_Unité_de_débit_volumique(Unité_de_débit_volumique):
                        list=['cm/s','cm/min','cm/h','cm/d','dm/s','dm/min','dm/h','dm/d','m/s','m/min','m/h','m/d','ml/s','ml/min','ml/h','ml/d','l/s','l/min','l/h','l/d','hl/s','hl/min','hl/h','hl/d','Ml/s','Ml/min','Ml/h','Ml/d','af/s','af/min','af/h','af/d','cf/s','cf/min','cf/h','cf/d','fl oz/s (us)','fl oz/min (us)','fl oz/h (us)','fl oz/d (us)','gal/s (us)','gal/min (us)','gal/h (us)','gal/d (us)','Mgal/s (us)','Mgal/min (us)','Mgal/h (us)','Mgal/d (us)','bbl/s (us;liq.)','bbl/min (us;liq.)','bbl/h (us;liq.)','bbl/d (us;liq.)','bbl/s (us;beer)','bbl/min (us;beer)','bbl/h (us;beer)','bbl/d (us;beer)','bbl/s (us;oil)','bbl/min (us;oil)','bbl/h (us;oil)','bbl/d (us;oil)','bbl/s (us;tank)','bbl/min (us;tank)','bbl/h (us;tank)','bbl/d (us;tank)','gal/s (imp)','gal/min (imp)','gal/h (imp)','gal/d (imp)','Mgal/s (imp)','Mgal/min (imp)','Mgal/h (imp)','Mgal/d (imp)','bbl/s (imp;oil)','bbl/min (imp;oil)','bbl/h (imp;oil)','bbl/d (imp;oil)','User vol./s','User vol./min','User vol./h','User vol./d','kgal/s (us)','kgal/min (us)','kgal/h (us)','kgal/d (us)'] 
                        return(list.index(Unité_de_débit_volumique))
                    def traduction_ordi_to_capt_Unité_de_masse(Unité_de_volume):
                        list=['cm³','dm³','m³','ml³','l','hl','Ml Mega','af','cf','fl oz (us)','gal (us)','Mgal (us)','bbl (us)','bbl (us;liq.)','bbl (us;beer)','bbl (us;oil)','bbl (us;tank)','gal (imp)','Mgal (imp)','bbl (imp;oil)','User vol','kgal']
                        return(list.index(Unité_de_volume))
                    def traduction_ordi_to_capt_Unité_du_débit_volumique_corrigé(Unité_du_débit_volumique_corrigé):
                        list=['Nl/s','Nl/min','Nl/h','Nl/d','Nm³/s','Nm³/min','Nm³/h','Nm³/d','Sm³/s','Sm³/min','Sm³/h','Sm³/d','Scf/s','Scf/min','Scf/h','Scf/d','Sgal/s (us)','Sgal/min (us)','Sgal/h (us)','Sgal/d (us)','Sbbl/s (us;liq.)','Sbbl/min (us;liq.)','Sbbl/h (us;liq.)','Sbbl/d (us;liq.)','Sgal/s (imp)','Sgal/min (imp)','Sgal/h (imp)','Sgal/d (imp)','User vol./s','User vol/min','User vol/h','User vol/d']
                        return(list.index(Unité_du_débit_volumique_corrigé))
                    def traduction_ordi_to_capt_Unité_de_volume_corrigé(Unité_de_volume_corrigé):
                        list=['Nl','Nm','Sm','Scf','Sl','Sgal (us)','Sbbl (us;liq.)','Sgal (imp)','UserCrVol.']
                        return(list.index(Unité_de_volume_corrigé))
                    def traduction_ordi_to_capt_Unité_de_densité(Unité_de_densité):
                        list=['g/cm³','kg/dm³','kg/l','kg/m³','SD4°C','SD15°C','SD20°C','SG4°C','SG15°C','SG20°C 1','lb/cf 1','lb/gal (us) 1','lb/bbl (us;liq.) 1','lb/bbl (us;beer) 1','lb/bbl (us;oil) 1','lb/bbl (us;tank) 1','lb/gal (imp) 1','lb/bbl (imp;beer) 1','lb/bbl (imp;oil) 2','User dens. 2','g/m 2'] 
                        return(list.index(Unité_de_densité))
                    def traduction_ordi_to_capt_Unité_de_densité_de_référence(Unité_de_densité_de_référence):
                        list=['g/Scm³','kg/Nl','kg/Nm³','kg/Sm³','lb/Scf']
                        return(list.index(Unité_de_densité_de_référence))
                    def traduction_ordi_to_capt_Unité_de_température(Unité_de_température):
                        list=['°C','K','°F','°R']
                        return(list.index(Unité_de_température))
                    def traduction_ordi_to_capt_Unité_de_pression(Unité_de_pression):
                        list=['bar a','psi a','bar g','psi g','Pa a','kPa a','MPa a','Pa g','kPa g','MPa g','User pres']
                        return(list.index(Unité_de_pression))
                    def traduction_ordi_to_capt_Format_date_heure(Format_date_heure):
                        list=['dd.mm.yy hh:mm','mm/dd/yy hh:mm am/pm','dd.mm.yy hh:mm am/pm','mm/dd/yy hh:mm']
                        return(list.index(Format_date_heure))
                    def traduction_ordi_to_capt_Affecter_variable_process_Suppression_débit_de_fuite(Affecter_variable_process_Suppression_débit_de_fuite):
                        list=['Off','Mass flow','Volume flow','Corrected volume flow']
                        return(list.index(Affecter_variable_process_Suppression_débit_de_fuite))

                    def traduction_ordi_to_capt_Affecter_variable_process_Détection_tube_partiellement_rempli(Affecter_variable_process_Détection_tube_partiellement_rempli):
                        if Affecter_variable_process_Détection_tube_partiellement_rempli =="off":
                            return(0)
                        elif Affecter_variable_process_Détection_tube_partiellement_rempli =="Density":
                            return(4)
                        elif Affecter_variable_process_Détection_tube_partiellement_rempli =="Reference density":
                            return(5)
                        
                        return(list.index(Affecter_variable_process_Détection_tube_partiellement_rempli))
                    def traduction_ordi_to_capt_Sélectionner_fluide(Sélectionner_fluide):
                        list=['Liquid', 'Gas']
                        return(list.index(Sélectionner_fluide))

                    def traduction_ordi_to_capt_Compensation_de_pression(Compensation_de_pression):
                        list=['Off', 'External value','External value']
                        list=['Off', 'Fixed value','External value']
                        return(list.index(Compensation_de_pression))

                    def traduction_ordi_to_capt_Mode_de_température(Mode_de_température):
                        list=['Internal measured value', 'External value']
                        return(list.index(Mode_de_température))
                        
                    def traduction_ordi_to_capt_Compensation_de_pression(Compensation_de_pression):
                        list=['Calculated reference density', 'Fixed reference density', 'External reference density', 'Reference density by API table 53']
                        return(list.index(Compensation_de_pression))

                    def traduction_ordi_to_capt_Commande_d_ajustage_du_zéro(Commande_d_ajustage_du_zéro):
                        if Commande_d_ajustage_du_zéro =="Flow in arrow direction":
                            return(0)
                        elif Commande_d_ajustage_du_zéro =="Flow against arrow direction":
                            return(1)

                    def traduction_ordi_to_capt_Sens_de_montage(Sens_de_montage):
                        if Sens_de_montage =="Cancel":
                            return(0)
                        elif Sens_de_montage =="Start":
                            return(1)
                        elif Sens_de_montage =="Zero point adjust failure":
                            return(2)
                        elif Sens_de_montage =="Busy":
                            return(8)
                        

                    def traduction_ordi_to_capt_Baudrate(Baudrate):
                        list=["1200 BAUD","2400 BAUD","4800 BAUD","9600 BAUD","19200 BAUD","38400 BAUD","57600 BAUD","115200 BAUD"]
                        return(list.index(Baudrate))
                    def traduction_ordi_to_capt_Mode_de_transfert_de_données(Mode_de_transfert_de_données):
                        list=["RTU","ASCII"]
                        return(list.index(Mode_de_transfert_de_données))
                    def traduction_ordi_to_capt_Parité(Parité):
                        list=["Even","Odd","None / 2 stop bits","None / 1 stop bit"]
                        return(list.index(Parité))
                    def traduction_ordi_to_capt_Ordre_des_octets(Ordre_des_octets):
                        list=["0-1-2-3","3-2-1-0","2-3-0-1","1-0-3-2"]
                        return(list.index(Ordre_des_octets))
                    def traduction_ordi_to_capt_Affecter_niveau_diagnostic(Affecter_niveau_diagnostic):
                        list=["Off","Warning","Alarm","Alarm or warning"]
                        return(list.index(Affecter_niveau_diagnostic))
                    def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                        list=["NaN value","Last valid value","Alarm","Alarm or warning"]
                        return(list.index(Mode_défaut))
                    def traduction_ordi_to_capt_Mode_interpréteur(Mode_interpréteur):
                        list=["Standard","Ignore surplus bytes"]
                        return(list.index(Mode_interpréteur))


                    def traduction_ordi_to_capt_RAZ_tous_les_totalisateurs(RAZ_tous_les_totalisateurs):
                        list=["Cancel","Reset + totalize"]
                        return(list.index(RAZ_tous_les_totalisateurs))
                    def traduction_ordi_to_capt_Affecter_variable_process_1(Affecter_variable_process_1):
                        list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                        return(list.index(Affecter_variable_process_1))
                    def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                        list=['g','kg','t','oz','lb','STon','User mass']
                        return(list.index(Unité_de_masse))
                    def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                        list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                        return(list.index(Mode_de_fonctionnement_totalisateur))
                    def traduction_ordi_to_capt_Contrôle_totalisateur_1(Contrôle_totalisateur_1):
                        list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                        return(list.index(Contrôle_totalisateur_1))
                    def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                        list=["Stop","Actual value","Last valid value"]
                        return(list.index(Mode_défaut))
                    def traduction_ordi_to_capt_Affecter_variable_process_2(Affecter_variable_process_2):
                        list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                        return(list.index(Affecter_variable_process_2))
                    def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                        list=['g','kg','t','oz','lb','STon','User mass']
                        return(list.index(Unité_de_masse))
                    def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                        list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                        return(list.index(Mode_de_fonctionnement_totalisateur))
                    def traduction_ordi_to_capt_Contrôle_totalisateur_2(Contrôle_totalisateur_2):
                        list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                        return(list.index(Contrôle_totalisateur_2))
                    def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                        list=["Stop","Actual value","Last valid value"]
                        return(list.index(Mode_défaut))

                    def traduction_ordi_to_capt_Affecter_variable_process_3(Affecter_variable_process_3):
                        list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                        return(list.index(Affecter_variable_process_3))
                    def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                        list=['g','kg','t','oz','lb','STon','User mass']
                        return(list.index(Unité_de_masse))
                    def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                        list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                        return(list.index(Mode_de_fonctionnement_totalisateur))
                    def traduction_ordi_to_capt_Contrôle_totalisateur_3(Contrôle_totalisateur_3):
                        list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                        return(list.index(Contrôle_totalisateur_3))
                    def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                        list=["Stop","Actual value","Last valid value"]
                        return(list.index(Mode_défaut))


                    if registre==2756 or registre==2755 or registre==2080 or registre==2021 or registre==2754 or registre==2022 or registre==2023 or registre==2758 or registre==2761 or registre==2760 or registre==2759 or registre==2757 or registre==2753 or registre==2081 or registre==2020:
                        print("in if write in 46")
                        
                        
                        try:
                            value=traduction_ordi_to_capt_Affecter_Numéro_de_diagnostic(rawvalue_a_envoyer)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()
                            window_erreur
                            window_erreur.mainloop()

                    if registre==6816 :
                        
                        try:
                            value=window_rapport.traduction_ordi_to_capt_Device_reset(registre)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()
                            window_erreur
                            window_erreur.mainloop()


                    if registre==6906 :
                        
                        try:  
                            value=window_rapport.traduction_ordi_to_capt_Sauvegarde_permanente(registre)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()
                            window_erreur
                            window_erreur.mainloop()
                    
                    if registre== 2100:
                        
                        try:
                                    
                            value=window_rapport.traduction_ordi_to_capt_Unité_de_débit_massique(registre)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()        
                            window_erreur
                            window_erreur.mainloop()
                        
                    if registre== 2101:
                        
                        try:
                                    
                            value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()        
                            window_erreur
                            window_erreur.mainloop()
                        
                    if registre== 2102:
                        
                        try:
                                    
                            value=window_rapport.traduction_ordi_to_capt_Unité_de_débit_volumique(registre)
                            self.communication.write_register_sensor(registre,value)
                                    
                        except:
                            window_erreur=window_erreur()        
                            window_erreur
                            window_erreur.mainloop()
                        
                        if registre== 2103:
                            
                            try:
                                        
                                value=window_rapport.traduction_capt_to_ordi_Unité_de_volume(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2104:
                            
                            try:
                                        
                                value=window_rapport.traduction_capt_to_ordi_Unité_du_débit_volumique_corrigé(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2105:
                            
                            try:
                                        
                                value=window_rapport.traduction_capt_to_ordi_Unité_de_volume_corrigé(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2106:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_densité(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2107:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_densité_de_référence(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2108:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_température(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                    
                        if registre== 2129:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_pression(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2149:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Format_date_heure(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 6816:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Device_reset(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2901:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Aperçu_des_options_logiciels(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                                
                        if registre== 6906:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Sauvegarde_permanente(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2100:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_débit_massique(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            

                        if registre== 2101:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2102:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_débit_volumique(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            

                        if registre== 2103:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2104:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_du_débit_volumique_corrigé(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                                
                        if registre== 2105:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_volume_corrigé(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2106:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_densité(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2107:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_densité_de_référence(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2108:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_température(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2129:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_pression(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2149:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Format_date_heure(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            

                        if registre== 5502:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Dépassement_débit(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            

                        if registre== 5100:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Affecter_variable_process_Suppression_débit_de_fuite(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5105:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Affecter_variable_process_Détection_tube_partiellement_rempli(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2441:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Sélectionner_fluide(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5228:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Compensation_de_pression(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2441:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_de_température(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5183:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Compensation_de_pression(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5514:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Commande_d_ajustage_du_zéro(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5128:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Sens_de_montage(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 5500:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Dépassement_débit(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4911:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Dépassement_débit(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4912:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Parité(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4913:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Ordre_des_octets(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4914:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Affecter_niveau_diagnostic(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4919:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_défaut(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 4924:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_interpréteur(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2608:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_RAZ_tous_les_totalisateurs(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2600:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Affecter_variable_process_1(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2601:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2604:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2607:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Contrôle_totalisateur_1(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                            

                        if registre== 2605:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_défaut(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                            
                                
                        if registre== 2800:
                            
                            try:
                                        
                                value=window_rapport.traduction_capt_to_ordi_Affecter_variable_process_2(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2801:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 2804:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                            
                        if registre== 2807:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Contrôle_totalisateur_2(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                            
                        if registre== 2805:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_défaut(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 3000:
                            
                            try:
                                        
                                value=window_rapport.traduction_capt_to_ordi_Affecter_variable_process_3(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 3001:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Unité_de_masse(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                                
                                
                        if registre== 3004:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 3007:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Contrôle_totalisateur_3(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        if registre== 3005:
                            
                            try:
                                        
                                value=window_rapport.traduction_ordi_to_capt_Mode_défaut(registre)
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()
                            
                        else:
                            
                            try:
                                self.communication.write_register_sensor(registre,value)
                                        
                            except:
                                window_erreur=window_erreur()        
                                window_erreur
                                window_erreur.mainloop()


                    attempt += 1
                except filelock.Timeout:
                    attempt += 1
                    print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
                    if attempt < max_attempts:
                        time.sleep(1)  # Optional: Add a delay before retrying

                if attempt == max_attempts:
                    print("Maximum attempts reached. Lock could not be acquired.")



        label=customtkinter.CTkLabel(self, text= Label ,compound="left", padx=5, anchor="w")
        liste_a_choix = customtkinter.CTkComboBox(self,values=liste_des_choix)
        button = customtkinter.CTkButton(self, text="Confirmé",command=lambda:write_new_value(registre,liste_a_choix.get()), width=100, height=24,)

        label.grid(row=len(self.label_list), column=0,padx=5, pady=(0, 10), sticky="w")
        liste_a_choix.grid(row=len(self.label_list), column=1, padx=5, pady=(0, 10), sticky="nsew")
        button.grid(row=len(self.label_list), column=2,padx=5, pady=(0, 10), sticky="nsew")

        self.label_list.append(label)
        self.nameEntry_list.append(liste_a_choix)
        self.button_list.append(button)
        liste_a_choix.set(valeur_du_capteur)    

    def modify_item(self,new_name_entry, item):
        for label,nameEntry_list, button in zip(self.label_list,self.nameEntry_list, self.button_list):
            if item == label.cget("text"):
                self.nameEntry_list.replace(nameEntry_list,new_name_entry)
                return
                """
                label.destroy()
                button.destroy()
                self.label_list.remove(label)
                self.button_list.remove(button)
                """
                
    #traduire les valeurs numérique en string
    #def traduction_numérique_to_string(self,valeur_bits): 
    #def traduction_string_to_numérique(self,valeur_bits):     
 
class window_rapport(customtkinter.CTk):
    def __init__(self,address,utilisateur):
        super().__init__()
        self.utilisateur=utilisateur


        file_path_Etat_Communication = 'fichier_excel/Etat Communication.xlsx'
        file_path_consommation = 'fichier_excel/consommation.xlsx'
        sheet_name = "Feuille1"


        def lock_excel_file():
            # Create a file lock
            file_lock = filelock.FileLock(file_path_Etat_Communication + ".lock")
            print("here def lock_excel")
            max_attempts = 1  # Maximum attempts to acquire the lock
            attempt = 0
            
            while attempt < max_attempts:
                try:
                    print("here while attempt first")
                    # Acquire the lock with a timeout of 20 seconds
                    with file_lock.acquire(timeout=20):
                        print(f"Lock acquired for first {file_path_Etat_Communication}")
                        
                        # Perform operations on the Excel file
                        wb = openpyxl.load_workbook(file_path_Etat_Communication)
                        sheet = wb.active

                        try:
                            from recuperation_donne import recuperationdonne
                        except:
                            print("error importation recuperation_donne")

                        
                        recuperationdonne=recuperationdonne(address,"Admin")
                        recuperationdonne.read_rapport()
                        self.utilisateur=utilisateur
                        self.title("Rapport d'appareil  ")
                        #self.grid_rowconfigure(0, weight=1)
                        self.columnconfigure(0, weight=1)
                        width= self.winfo_screenwidth()
                        height= self.winfo_screenheight()
                        #current_dir="image projet\icone_rapport_d'appareil.png"
                        self.scrollable_label_button_frame = ScrollableLabelButtonFrame(master=self, width=width,height=height,communication=recuperationdonne,utilisateur=self.utilisateur, corner_radius=0)
                        self.scrollable_label_button_frame.grid(row=0, column=0, padx=5, pady=0, sticky="nsew")
                                   
                        self.geometry("%dx%d"% (width, height))
                        #expert
                        self.scrollable_label_button_frame.add_item(f"\t\t retour "," ")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t expert")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t expert")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t État verrouillage")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t Droits d'accès via logiciel ",recuperationdonne.Droits_d_accès_via_logiciel,2177)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t Entrer code d'accès ",recuperationdonne.Entrer_code_d_accès,2176)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t expert")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Traitement événement")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t  Traitement événement")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Temporisation alarme ",recuperationdonne.Temporisation_alarme,6807,1)

                        def traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(Affecter_Numéro_de_diagnostic):
                            if Affecter_Numéro_de_diagnostic ==0:
                                return("Off")
                            elif Affecter_Numéro_de_diagnostic ==1:
                                return("Logbook entry only")
                            elif Affecter_Numéro_de_diagnostic ==2:
                                return("Warning")
                            elif Affecter_Numéro_de_diagnostic ==3:
                                return("Alarm")

                        def traduction_ordi_to_capt_Affecter_Numéro_de_diagnostic(Affecter_Numéro_de_diagnostic):
                            if Affecter_Numéro_de_diagnostic =="off":
                                return(0)
                            elif Affecter_Numéro_de_diagnostic =="Logbook entry only":
                                return(1)
                            elif Affecter_Numéro_de_diagnostic =="Warning":
                                return(2)
                            elif Affecter_Numéro_de_diagnostic =="Alarm":
                                return(3)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Comportement du diagnostic")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t  Comportement du diagnostic")
                        #self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.140",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(),["off","Logbook entry only","Warning","Alarm"])
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.046",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_046),["off","Logbook entry only","Warning","Alarm"],2755)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.144",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_144),["off","Logbook entry only","Warning","Alarm"],2080)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.832",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_832),["off","Logbook entry only","Warning","Alarm"],2758)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.833",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_833),["off","Logbook entry only","Warning","Alarm"],2761)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.834",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_834),["off","Logbook entry only","Warning","Alarm"],2760)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.835",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_835),["off","Logbook entry only","Warning","Alarm"],2759)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.912",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_912),["off","Logbook entry only","Warning","Alarm"],2757)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.913",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_913),["off","Logbook entry only","Warning","Alarm"],2753)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.944",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_944),["off","Logbook entry only","Warning","Alarm"],2081)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.192",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_192),["off","Logbook entry only","Warning","Alarm"],2021)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.274",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_274),["off","Logbook entry only","Warning","Alarm"],2754)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.392",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_392),["off","Logbook entry only","Warning","Alarm"],2022)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.592",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_592),["off","Logbook entry only","Warning","Alarm"],2023)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t\t Affecter Numéro de diagnostic no.992",traduction_capt_to_ordi_Affecter_Numéro_de_diagnostic(recuperationdonne.Affecter_Numéro_de_diagnostic_992),["off","Logbook entry only","Warning","Alarm"],2020)

                        def traduction_capt_to_ordi_Device_reset(Device_reset):
                            if Device_reset ==0:
                                return("Cancel")
                            elif Device_reset ==1:
                                return("Restart device")
                            elif Device_reset ==2:
                                return("To delivery settings")
                            

                        def traduction_ordi_to_capt_Device_reset(Device_reset):
                            if Device_reset =="Cancel":
                                return(0)
                            elif Device_reset =="Restart device":
                                return(1)
                            elif Device_reset =="To delivery settings":
                                return(2)
                            
                        def traduction_capt_to_ordi_Aperçu_des_options_logiciels(Aperçu_des_options_logiciels):
                            if Aperçu_des_options_logiciels ==16384:
                                return("Heartbeat Monitoring")
                            elif Aperçu_des_options_logiciels ==4:
                                return("Concentration")
                            elif Aperçu_des_options_logiciels ==32768:
                                return("Heartbeat Verification")
                            elif Aperçu_des_options_logiciels ==64:
                                return("Viscosity")

                        def traduction_ordi_to_capt_Aperçu_des_options_logiciels(Aperçu_des_options_logiciels):
                            if Aperçu_des_options_logiciels =="Heartbeat Monitoring":
                                return(16384)
                            elif Aperçu_des_options_logiciels =="Concentration":
                                return(4)
                            elif Aperçu_des_options_logiciels =="Heartbeat Verification":
                                return(32768)
                            elif Aperçu_des_options_logiciels =="Viscosity":
                                return(64)
                            
                        def traduction_capt_to_ordi_Sauvegarde_permanente(Sauvegarde_permanente):
                            if Sauvegarde_permanente ==0:
                                return("Off")
                            elif Sauvegarde_permanente ==1:
                                return("On")
                            

                        def traduction_ordi_to_capt_Sauvegarde_permanente(Sauvegarde_permanente):
                            if Sauvegarde_permanente =="Off":
                                return(0)
                            elif Sauvegarde_permanente =="On":
                                return(1)
                        print(traduction_capt_to_ordi_Aperçu_des_options_logiciels(recuperationdonne.Aperçu_des_options_logiciels))
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Administration")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t  Administration")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Reset appareil",traduction_capt_to_ordi_Device_reset(recuperationdonne.Reset_appareil),["Cancel","Restart device","To delivery setting"],6816)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Activer options software ",recuperationdonne.Activer_options_software,2794)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Sauvegarde permanente",traduction_capt_to_ordi_Sauvegarde_permanente(recuperationdonne.Sauvegarde_permanente),["off","on"],6906)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Désignation du point de mesure  ",recuperationdonne.)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Valeur mesurée")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Variables process")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t  Variables process")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Débit massique  ",recuperationdonne.Débit_massique,2006,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Débit volumique  ",recuperationdonne.Débit_volumique,2008,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Débit volumique corrigé ",recuperationdonne.Débit_volumique_corrigé,2010,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Densité  ",recuperationdonne.Densité,2012,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Densité de référence ",recuperationdonne.Densité_de_référence,2014,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Température ",recuperationdonne.Température,2016,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Valeur de pression ",recuperationdonne.Valeur_de_pression,2088,1)
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Totalisateur")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t  Totalisateur")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur totalisateur 1 ",recuperationdonne.Valeur_totalisateur_1,2609,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Dépassement totalisateur 1 ",recuperationdonne.Dépassement_totalisateur_1,2611,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur totalisateur 2  ",recuperationdonne.Valeur_totalisateur_2,2809,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Dépassement totalisateur 2  ",recuperationdonne.Dépassement_totalisateur_2,2811,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur totalisateur 3 ",recuperationdonne.Valeur_totalisateur_3,3009,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Dépassement totalisateur 3 ",recuperationdonne.Dépassement_totalisateur_3,3011,1)

                        def traduction_capt_to_ordi_Unité_de_débit_massique(Unité_de_débit_massique):
                            list=['g/s','g/min','g/h','g/d','kg/s','kg/min','kg/h','kg/d','t/s','t/min','t/h','t/d','oz/s','oz/min','oz/h','oz/d','lb/s','lb/min','lb/h','lb/d','STon/s','STon/min','STon/h','STon/d','User mass/s','User mass/min','User mass/h']  
                            return(list[Unité_de_débit_massique])
                            

                        def traduction_ordi_to_capt_Unité_de_débit_massique(Unité_de_débit_massique):
                            list=['g/s','g/min','g/h','g/d','kg/s','kg/min','kg/h','kg/d','t/s','t/min','t/h','t/d','oz/s','oz/min','oz/h','oz/d','lb/s','lb/min','lb/h','lb/d','STon/s','STon/min','STon/h','STon/d','User mass/s','User mass/min','User mass/h']  
                            return(list.index(Unité_de_débit_massique))
                            
                        def traduction_capt_to_ordi_Unité_de_masse(Unité_de_masse):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list[Unité_de_masse])
                            

                        def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list.index(Unité_de_masse))


                        def traduction_capt_to_ordi_Unité_de_débit_volumique(Unité_de_débit_volumique):
                            list=['cm³/s','cm³/min','cm³/h','cm³/d','dm³/s','dm³/min','dm³/h','dm³/d','m³/s','m³/min','m³/h','m³/d','ml/s','ml/min','ml/h','ml/d','l/s','l/min','l/h','l/d','hl/s','hl/min','hl/h','hl/d','Ml/s','Ml/min','Ml/h','Ml/d','af/s','af/min','af/h','af/d','cf/s','cf/min','cf/h','cf/d','fl oz/s (us)','fl oz/min (us)','fl oz/h (us)','fl oz/d (us)','gal/s (us)','gal/min (us)','gal/h (us)','gal/d (us)','Mgal/s (us)','Mgal/min (us)','Mgal/h (us)','Mgal/d (us)','bbl/s (us;liq.)','bbl/min (us;liq.)','bbl/h (us;liq.)','bbl/d (us;liq.)','bbl/s (us;beer)','bbl/min (us;beer)','bbl/h (us;beer)','bbl/d (us;beer)','bbl/s (us;oil)','bbl/min (us;oil)','bbl/h (us;oil)','bbl/d (us;oil)','bbl/s (us;tank)','bbl/min (us;tank)','bbl/h (us;tank)','bbl/d (us;tank)','gal/s (imp)','gal/min (imp)','gal/h (imp)','gal/d (imp)','Mgal/s (imp)','Mgal/min (imp)','Mgal/h (imp)','Mgal/d (imp)','bbl/s (imp;oil)','bbl/min (imp;oil)','bbl/h (imp;oil)','bbl/d (imp;oil)','User vol./s','User vol./min','User vol./h','User vol./d','kgal/s (us)','kgal/min (us)','kgal/h (us)','kgal/d (us)'] 
                            return(list[Unité_de_débit_volumique])
                            

                        def traduction_ordi_to_capt_Unité_de_débit_volumique(Unité_de_débit_volumique):
                            list=['cm/s','cm/min','cm/h','cm/d','dm/s','dm/min','dm/h','dm/d','m/s','m/min','m/h','m/d','ml/s','ml/min','ml/h','ml/d','l/s','l/min','l/h','l/d','hl/s','hl/min','hl/h','hl/d','Ml/s','Ml/min','Ml/h','Ml/d','af/s','af/min','af/h','af/d','cf/s','cf/min','cf/h','cf/d','fl oz/s (us)','fl oz/min (us)','fl oz/h (us)','fl oz/d (us)','gal/s (us)','gal/min (us)','gal/h (us)','gal/d (us)','Mgal/s (us)','Mgal/min (us)','Mgal/h (us)','Mgal/d (us)','bbl/s (us;liq.)','bbl/min (us;liq.)','bbl/h (us;liq.)','bbl/d (us;liq.)','bbl/s (us;beer)','bbl/min (us;beer)','bbl/h (us;beer)','bbl/d (us;beer)','bbl/s (us;oil)','bbl/min (us;oil)','bbl/h (us;oil)','bbl/d (us;oil)','bbl/s (us;tank)','bbl/min (us;tank)','bbl/h (us;tank)','bbl/d (us;tank)','gal/s (imp)','gal/min (imp)','gal/h (imp)','gal/d (imp)','Mgal/s (imp)','Mgal/min (imp)','Mgal/h (imp)','Mgal/d (imp)','bbl/s (imp;oil)','bbl/min (imp;oil)','bbl/h (imp;oil)','bbl/d (imp;oil)','User vol./s','User vol./min','User vol./h','User vol./d','kgal/s (us)','kgal/min (us)','kgal/h (us)','kgal/d (us)'] 
                            return(list.index(Unité_de_débit_volumique))
                        

                        def traduction_capt_to_ordi_Unité_de_volume(Unité_de_volume):
                            list=['cm³','dm³','m³','ml³','l','hl','Ml Mega','af','cf','fl oz (us)','gal (us)','Mgal (us)','bbl (us)','bbl (us;liq.)','bbl (us;beer)','bbl (us;oil)','bbl (us;tank)','gal (imp)','Mgal (imp)','bbl (imp;oil)','User vol','kgal']
                            return(list[Unité_de_volume])
                            

                        def traduction_ordi_to_capt_Unité_de_masse(Unité_de_volume):
                            list=['cm³','dm³','m³','ml³','l','hl','Ml Mega','af','cf','fl oz (us)','gal (us)','Mgal (us)','bbl (us)','bbl (us;liq.)','bbl (us;beer)','bbl (us;oil)','bbl (us;tank)','gal (imp)','Mgal (imp)','bbl (imp;oil)','User vol','kgal']
                            return(list.index(Unité_de_volume))
                        

                        def traduction_capt_to_ordi_Unité_du_débit_volumique_corrigé(Unité_du_débit_volumique_corrigé):
                            list=['Nl/s','Nl/min','Nl/h','Nl/d','Nm³/s','Nm³/min','Nm³/h','Nm³/d','Sm³/s','Sm³/min','Sm³/h','Sm³/d','Scf/s','Scf/min','Scf/h','Scf/d','Sgal/s (us)','Sgal/min (us)','Sgal/h (us)','Sgal/d (us)','Sbbl/s (us;liq.)','Sbbl/min (us;liq.)','Sbbl/h (us;liq.)','Sbbl/d (us;liq.)','Sgal/s (imp)','Sgal/min (imp)','Sgal/h (imp)','Sgal/d (imp)','User vol./s','User vol/min','User vol/h','User vol/d']
                            return(list[Unité_du_débit_volumique_corrigé])
                            

                        def traduction_ordi_to_capt_Unité_du_débit_volumique_corrigé(Unité_du_débit_volumique_corrigé):
                            list=['Nl/s','Nl/min','Nl/h','Nl/d','Nm³/s','Nm³/min','Nm³/h','Nm³/d','Sm³/s','Sm³/min','Sm³/h','Sm³/d','Scf/s','Scf/min','Scf/h','Scf/d','Sgal/s (us)','Sgal/min (us)','Sgal/h (us)','Sgal/d (us)','Sbbl/s (us;liq.)','Sbbl/min (us;liq.)','Sbbl/h (us;liq.)','Sbbl/d (us;liq.)','Sgal/s (imp)','Sgal/min (imp)','Sgal/h (imp)','Sgal/d (imp)','User vol./s','User vol/min','User vol/h','User vol/d']
                            return(list.index(Unité_du_débit_volumique_corrigé))


                        def traduction_capt_to_ordi_Unité_de_volume_corrigé(Unité_de_volume_corrigé):
                            list=['Nl','Nm','Sm','Scf','Sl','Sgal (us)','Sbbl (us;liq.)','Sgal (imp)','UserCrVol.']
                            return(list[Unité_de_volume_corrigé])
                            

                        def traduction_ordi_to_capt_Unité_de_volume_corrigé(Unité_de_volume_corrigé):
                            list=['Nl','Nm','Sm','Scf','Sl','Sgal (us)','Sbbl (us;liq.)','Sgal (imp)','UserCrVol.']
                            return(list.index(Unité_de_volume_corrigé))
                        

                        def traduction_capt_to_ordi_Unité_de_densité(Unité_de_densité):
                            list=['g/cm³','kg/dm³','kg/l','kg/m³','SD4°C','SD15°C','SD20°C','SG4°C','SG15°C','SG20°C 1','lb/cf 1','lb/gal (us) 1','lb/bbl (us;liq.) 1','lb/bbl (us;beer) 1','lb/bbl (us;oil) 1','lb/bbl (us;tank) 1','lb/gal (imp) 1','lb/bbl (imp;beer) 1','lb/bbl (imp;oil) 2','User dens. 2','g/m 2']  
                            return(list[Unité_de_densité])
                            

                        def traduction_ordi_to_capt_Unité_de_densité(Unité_de_densité):
                            list=['g/cm³','kg/dm³','kg/l','kg/m³','SD4°C','SD15°C','SD20°C','SG4°C','SG15°C','SG20°C 1','lb/cf 1','lb/gal (us) 1','lb/bbl (us;liq.) 1','lb/bbl (us;beer) 1','lb/bbl (us;oil) 1','lb/bbl (us;tank) 1','lb/gal (imp) 1','lb/bbl (imp;beer) 1','lb/bbl (imp;oil) 2','User dens. 2','g/m 2'] 
                            return(list.index(Unité_de_densité))
                        
                        

                        def traduction_capt_to_ordi_Unité_de_densité_de_référence(Unité_de_densité_de_référence):
                            list=['g/Scm³','kg/Nl','kg/Nm³','kg/Sm³','lb/Scf']
                            return(list[Unité_de_densité_de_référence])
                            

                        def traduction_ordi_to_capt_Unité_de_densité_de_référence(Unité_de_densité_de_référence):
                            list=['g/Scm³','kg/Nl','kg/Nm³','kg/Sm³','lb/Scf']
                            return(list.index(Unité_de_densité_de_référence))


                        def traduction_capt_to_ordi_Unité_de_température(Unité_de_température):
                            list=['°C','K','°F','°R']
                            return(list[Unité_de_température])
                            

                        def traduction_ordi_to_capt_Unité_de_température(Unité_de_température):
                            list=['°C','K','°F','°R']
                            return(list.index(Unité_de_température))
                        

                        def traduction_capt_to_ordi_Unité_de_pression(Unité_de_pression):
                            list=['bar a','psi a','bar g','psi g','Pa a','kPa a','MPa a','Pa g','kPa g','MPa g','User pres']
                            return(list[Unité_de_pression])
                            

                        def traduction_ordi_to_capt_Unité_de_pression(Unité_de_pression):
                            list=['bar a','psi a','bar g','psi g','Pa a','kPa a','MPa a','Pa g','kPa g','MPa g','User pres']
                            return(list.index(Unité_de_pression))
                        

                        def traduction_capt_to_ordi_Format_date_heure(Format_date_heure):
                            list=['dd.mm.yy hh:mm','mm/dd/yy hh:mm am/pm','dd.mm.yy hh:mm am/pm','mm/dd/yy hh:mm']
                            return(list[Format_date_heure])
                            

                        def traduction_ordi_to_capt_Format_date_heure(Format_date_heure):
                            list=['dd.mm.yy hh:mm','mm/dd/yy hh:mm am/pm','dd.mm.yy hh:mm am/pm','mm/dd/yy hh:mm']
                            return(list.index(Format_date_heure))


                                                                                                                                        
                                                                                                                                        
                                                                                                                                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Unités système")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Unités système")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de débit massique",traduction_capt_to_ordi_Unité_de_débit_massique(recuperationdonne.Unité_de_débit_massique),["g/s","on"],2100)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de masse",traduction_capt_to_ordi_Unité_de_masse(recuperationdonne.Unité_de_masse),["g/s","on"],2101)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de débit volumique",traduction_capt_to_ordi_Unité_de_débit_volumique(recuperationdonne.Unité_de_débit_volumique),["g/s","on"],2102)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de volume",traduction_capt_to_ordi_Unité_de_volume(recuperationdonne.Unité_de_volume),["g/s","on"],2103)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité du débit volumique corrigé",traduction_capt_to_ordi_Unité_du_débit_volumique_corrigé(recuperationdonne.Unité_du_débit_volumique_corrigé),["g/s","on"],2104)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de volume corrigé ",traduction_capt_to_ordi_Unité_de_volume_corrigé(recuperationdonne.Unité_de_volume_corrigé),["g/s","on"],2105)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de densité",traduction_capt_to_ordi_Unité_de_densité(recuperationdonne.Unité_de_densité),["g/s","on"],2106)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de densité de référence ",traduction_capt_to_ordi_Unité_de_densité_de_référence(recuperationdonne.Unité_de_densité_de_référence),["g/s","on"],2107)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de température",traduction_capt_to_ordi_Unité_de_température(recuperationdonne.Unité_de_température),["g/s","on"],2108)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de pression  ",traduction_capt_to_ordi_Unité_de_pression(recuperationdonne.Unité_de_pression),["g/s","on"],2129)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Format date/heure",traduction_capt_to_ordi_Format_date_heure(recuperationdonne.Format_date_heure),["g/s","on"],2149)


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Unités spécifiques utilisateur")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Unités spécifiques utilisateur")
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Nom unité masse utilisateur  ",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Facteur masse utilisateur ",recuperationdonne.User_mass_factor,2114,1)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Nom unité volume utilisateur   ",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Facteur volume utilisateur ",recuperationdonne.User_volume_factor,2118,1)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Nom unité volume corrigé utilisateur  ",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Facteur volume corrigé utilisateur  ",recuperationdonne.User_corrected_volume_factor,2572,1)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Nom unité densité utilisateur   ",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Offset densité utilisateur  ",recuperationdonne.User_density_offset,2555,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Facteur densité utilisateur  ",recuperationdonne.User_density_factor,2122,1)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Texte pression utilisateur ",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Compensation de pression utilisateur   ",recuperationdonne.User_pressure_offset,2565,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Facteur de pression utilisateur  ",recuperationdonne.User_pressure_factor,2563,1)

                        def traduction_capt_to_ordi_Dépassement_débit(ordi_Dépassement_débit):
                            list=['Off','on']
                            return(list[ordi_Dépassement_débit])
                            

                        def traduction_ordi_to_capt_Dépassement_débit(ordi_Dépassement_débit):
                            list=['Off','on']
                            return(list.index(ordi_Dépassement_débit))
                        print(traduction_capt_to_ordi_Dépassement_débit(recuperationdonne.Dépassement_débit))
                        print("recuperationdonne.Dépassement_débit:",recuperationdonne.Dépassement_débit)
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Paramètres process")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Paramètres process")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortissement débit ",recuperationdonne.Amortissement_débit,5509,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortissement densité ",recuperationdonne.Amortissement_densité,5507,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortissement température ",recuperationdonne.Amortissement_température,5126,1)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Dépassement débit ",traduction_capt_to_ordi_Dépassement_débit(recuperationdonne.Dépassement_débit),["On","Off"],5502)

                        def traduction_capt_to_ordi_Affecter_variable_process_Suppression_débit_de_fuite(Affecter_variable_process_Suppression_débit_de_fuite):
                            list=['Off','Mass flow','Volume flow','Corrected volume flow']
                            return(list[Affecter_variable_process_Suppression_débit_de_fuite])
                            

                        def traduction_ordi_to_capt_Affecter_variable_process_Suppression_débit_de_fuite(Affecter_variable_process_Suppression_débit_de_fuite):
                            list=['Off','Mass flow','Volume flow','Corrected volume flow']
                            return(list.index(Affecter_variable_process_Suppression_débit_de_fuite))
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Suppression débit de fuite")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Suppression débit de fuite")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Affecter variable process",traduction_capt_to_ordi_Affecter_variable_process_Suppression_débit_de_fuite(recuperationdonne.Affecter_variable_process_Suppression_débit_de_fuite),["Off","Mass flow","Volume flow","Corrected volume flow"],5100)     
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur 'on'débit de fuite",recuperationdonne.Valeur_on_débit_de_fuite,5137,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur 'off'débit de fuite ",recuperationdonne.Valeur_off_débit_de_fuite,5103,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Suppression effet pulsatoire  ",recuperationdonne.Suppression_effet_pulsatoire,5139,1)


                        def traduction_capt_to_ordi_Affecter_variable_process_Détection_tube_partiellement_rempli(Affecter_variable_process_Détection_tube_partiellement_rempli):
                            if Affecter_variable_process_Détection_tube_partiellement_rempli ==0:
                                return("Off")
                            elif Affecter_variable_process_Détection_tube_partiellement_rempli ==4:
                                return("Density")
                            elif Affecter_variable_process_Détection_tube_partiellement_rempli ==5:
                                return("Reference density")
                            return(list[Affecter_variable_process_Détection_tube_partiellement_rempli])
                            

                        def traduction_ordi_to_capt_Affecter_variable_process_Détection_tube_partiellement_rempli(Affecter_variable_process_Détection_tube_partiellement_rempli):
                            if Affecter_variable_process_Détection_tube_partiellement_rempli =="off":
                                return(0)
                            elif Affecter_variable_process_Détection_tube_partiellement_rempli =="Density":
                                return(4)
                            elif Affecter_variable_process_Détection_tube_partiellement_rempli =="Reference density":
                                return(5)
                            
                            return(list.index(Affecter_variable_process_Détection_tube_partiellement_rempli))
                        

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Détection tube partiellement rempli")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Détection tube partiellement rempli")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Assign process variable",traduction_capt_to_ordi_Affecter_variable_process_Détection_tube_partiellement_rempli(recuperationdonne.Affecter_variable_process_Détection_tube_partiellement_rempli),["Off ","Density","Reference density"],5106)   
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur basse détect. tube part. rempli",recuperationdonne.Valeur_basse_détect_tube_part_rempli,5109,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur haute détect. tube part. rempli  ",recuperationdonne.Valeur_haute_détect_tube_part_rempl,5111,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Temps réponse détect. tube part. rempl  ",recuperationdonne.Tempsréponsedétect_tube_part_rempli,5107,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortis. max. détect. tube part. rempli  ",recuperationdonne.Amortis_max_détect_tube_part_rempli,2413,1)


                        def traduction_capt_to_ordi_Sélectionner_fluide(Sélectionner_fluide):
                            list=['Liquid', 'Gas']
                            return(list[Sélectionner_fluide])
                            

                        def traduction_ordi_to_capt_Sélectionner_fluide(Sélectionner_fluide):
                            list=['Liquid', 'Gas']
                            return(list.index(Sélectionner_fluide))

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Mode de mesure")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Mode de mesure")
                        
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Sélectionner fluide",traduction_capt_to_ordi_Affecter_variable_process_Détection_tube_partiellement_rempli(recuperationdonne.Sélectionner_fluide),['Liquide','Gas'],2441)

                        def traduction_capt_to_ordi_Compensation_de_pression(Compensation_de_pression):
                            list=['Off', 'External value','External value ']
                            return(list[Compensation_de_pression])
                            

                        def traduction_ordi_to_capt_Compensation_de_pression(Compensation_de_pression):
                            list=['Off', 'External value','External value']
                            list=['Off', 'Fixed value','External value']
                            return(list.index(Compensation_de_pression))
                        
                        def traduction_capt_to_ordi_Mode_de_température(Mode_de_température):
                            list=['Internal measured value', 'External value']
                            return(list[Mode_de_température])
                            

                        def traduction_ordi_to_capt_Mode_de_température(Mode_de_température):
                            list=['Internal measured value', 'External value']
                            return(list.index(Mode_de_température))

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Compensation externe")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Compensation externe")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Compensation de pression ",traduction_capt_to_ordi_Compensation_de_pression(recuperationdonne.Compensation_de_pression),["Off ","Fixed value","External value"],5183)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode de température  ",traduction_capt_to_ordi_Mode_de_température(recuperationdonne.Mode_de_température),["Internal measured value ","External value"],5514)
                        

                        def traduction_capt_to_ordi_Compensation_de_pression(Compensation_de_pression):
                            list=['Calculated reference density', 'Fixed reference density', 'External reference density', 'Reference density by API table 53']
                            return(list[Compensation_de_pression])
                            

                        def traduction_ordi_to_capt_Compensation_de_pression(Compensation_de_pression):
                            list=['Calculated reference density', 'Fixed reference density', 'External reference density', 'Reference density by API table 53']
                            return(list.index(Compensation_de_pression))


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Valeurs calculées")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Calcul du débit volumique corrigé")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Calcul du débit volumique corrigé")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Compensation de pression",traduction_capt_to_ordi_Compensation_de_pression(recuperationdonne.Compensation_de_pression),["Calculated reference densit","Fixed reference density","External reference density","Reference density by API table 53"],5183)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Température de référence     ",recuperationdonne.Température_de_référence,5184)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Coefficient de dilation linéaire     ",recuperationdonne.Coefficient_de_dilation_linéaire,5131,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Coefficient de dilatation au carré   ",recuperationdonne.Coefficient_de_dilatation_au_carré,5133,1)

                        def traduction_capt_to_ordi_Commande_d_ajustage_du_zéro(Commande_d_ajustage_du_zéro):
                            if Commande_d_ajustage_du_zéro ==0:
                                return("Flow in arrow direction")
                            elif Commande_d_ajustage_du_zéro ==1:
                                return("Flow against arrow direction")

                        def traduction_ordi_to_capt_Commande_d_ajustage_du_zéro(Commande_d_ajustage_du_zéro):
                            if Commande_d_ajustage_du_zéro =="Flow in arrow direction":
                                return(0)
                            elif Commande_d_ajustage_du_zéro =="Flow against arrow direction":
                                return(1)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Ajustage capteur")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Ajustage capteur")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Sens de montage  ",traduction_capt_to_ordi_Commande_d_ajustage_du_zéro(recuperationdonne.Sens_de_montage),["Flow in arrow direction","Flow against arrow direction"],5500)

                        def traduction_capt_to_ordi_Sens_de_montage(Sens_de_montage):
                            if Sens_de_montage ==0:
                                return("Cancel")
                            elif Sens_de_montage ==1:
                                return("Start")
                            elif Sens_de_montage ==2:
                                return("Zero point adjust failure")
                            elif Sens_de_montage ==8:
                                return("Busy")

                        def traduction_ordi_to_capt_Sens_de_montage(Sens_de_montage):
                            if Sens_de_montage =="Cancel":
                                return(0)
                            elif Sens_de_montage =="Start":
                                return(1)
                            elif Sens_de_montage =="Zero point adjust failure":
                                return(2)
                            elif Sens_de_montage =="Busy":
                                return(8)
                        


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Ajustage du zéro")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Ajustage du zéro")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Commande d'ajustage du zéro  ",traduction_capt_to_ordi_Sens_de_montage(recuperationdonne.Commande_d_ajustage_du_zéro),["Cancel   ","Start","Zero point adjust failure","Busy"],5120)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t En cours ",recuperationdonne.En_cours,6796)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Ajustage densité")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Ajustage variable process")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Ajustage variable process")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de débit massique  ",recuperationdonne.Offset_de_débit_massique,5520,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de débit massique ",recuperationdonne.Facteur_de_débit_massique,5518,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de débit volumique ",recuperationdonne.Offset_de_débit_volumique,5524,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de débit volumique    ",recuperationdonne.Facteur_de_débit_volumique,5522,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de densité ",recuperationdonne.Offset_de_densité,5528,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de densité    ",recuperationdonne.Facteur_de_densité,5526,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de débit volumique corrigé ",recuperationdonne.Offset_de_débit_volumique_corrigé,2043,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de débit volumique corrigé",recuperationdonne.Facteur_de_débit_volumique_corrigé,2075,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de densité de référence",recuperationdonne.Offset_de_densité_de_référence,2045,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de densité de référence   ",recuperationdonne.Facteur_de_densité_de_référence,2041,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Offset de température ",recuperationdonne.Offset_de_température,5532,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Facteur de température",recuperationdonne.Facteur_de_température,5530,1)


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Étalonnage")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Étalonnage")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Facteur d'étalonnage",recuperationdonne.Calibration_factor,7512,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Zéro ",recuperationdonne.Zero_point,7526,1)
                        #self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Diamètre nominal",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C0",recuperationdonne.C0,7500,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C1",recuperationdonne.C1,7502,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C2",recuperationdonne.C2,7504,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C3",recuperationdonne.C3,7506,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C4",recuperationdonne.C4,7508,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t C5",recuperationdonne.C5,7510,1)


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Points test")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Points test")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Fréquence d'oscillation 0 ",recuperationdonne.Fréquence_d_oscillation_0,9500,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Fréquence d'oscillation 1 ",recuperationdonne.Fréquence_d_oscillation_1,9502,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Fluctuations fréquence 0  ",recuperationdonne.Fluctuations_fréquence_0,2497,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Fluctuations fréquence 1  ",recuperationdonne.Fluctuations_fréquence_1,2499,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amplitude de l'oscillation 0  ",recuperationdonne.Amplitude_de_l_oscillation_0,2448,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amplitude de l'oscillation 1  ",recuperationdonne.Amplitude_de_l_oscillation_1,2450,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortissement de l'oscillation",recuperationdonne.Amortissement_de_l_oscillation_0,9504,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Amortissement de l'oscillation 1  ",recuperationdonne.Amortissement_de_l_oscillation_1,9506,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Fluctuations amortissement tube 0 ",recuperationdonne.Fluctuations_amortissement_tube_0,2501,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t  Fluctuations amortissement tube 1",recuperationdonne.Fluctuations_amortissement_tube_1,2503,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t symétrie signal",recuperationdonne.Asymétrie_signal,2442,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Température électroniqu",recuperationdonne.Température_électronique,2456,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Courant d'excitation 0 ",recuperationdonne.Courant_d_excitation_0,9508,1)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Courant d'excitation 1 ",recuperationdonne.Courant_d_excitation_1,9510,1)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Raw values")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t RawMassFlow  ",recuperationdonne.RawMassFlow,10231,1)
                        """
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Supervision")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Oscillation")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Limit value measuring tube damping",)
                        """
                        def traduction_capt_to_ordi_Baudrate(Baudrate):
                            list=["1200 BAUD","2400 BAUD","4800 BAUD","9600 BAUD","19200 BAUD","38400 BAUD","57600 BAUD","115200 BAUD"]
                            return(list[Baudrate])
                            

                        def traduction_ordi_to_capt_Baudrate(Baudrate):
                            list=["1200 BAUD","2400 BAUD","4800 BAUD","9600 BAUD","19200 BAUD","38400 BAUD","57600 BAUD","115200 BAUD"]
                            return(list.index(Baudrate))

                        def traduction_capt_to_ordi_Mode_de_transfert_de_données(Mode_de_transfert_de_données):
                            list=["RTU","ASCII  "]
                            return(list[Mode_de_transfert_de_données])
                        

                        def traduction_ordi_to_capt_Mode_de_transfert_de_données(Mode_de_transfert_de_données):
                            list=["RTU","ASCII"]
                            return(list.index(Mode_de_transfert_de_données))

                        def traduction_capt_to_ordi_Parité(Parité):
                            list=["Even","Odd","None / 2 stop bits","None / 1 stop bit"]
                            return(list[Parité])
                                

                        def traduction_ordi_to_capt_Parité(Parité):
                            list=["Even","Odd","None / 2 stop bits","None / 1 stop bit"]
                            return(list.index(Parité))
                        
                        def traduction_capt_to_ordi_Ordre_des_octets(Ordre_des_octets):
                            list=["0-1-2-3","3-2-1-0","2-3-0-1","1-0-3-2"]
                            return(list[Ordre_des_octets])
                                

                        def traduction_ordi_to_capt_Ordre_des_octets(Ordre_des_octets):
                            list=["0-1-2-3","3-2-1-0","2-3-0-1","1-0-3-2"]
                            return(list.index(Ordre_des_octets))
                        
                        def traduction_capt_to_ordi_Affecter_niveau_diagnostic (Affecter_niveau_diagnostic ):
                            list=["Off","Warning","Alarm","Alarm or warning"]
                            return(list[Affecter_niveau_diagnostic ])  

                        def traduction_ordi_to_capt_Affecter_niveau_diagnostic(Affecter_niveau_diagnostic):
                            list=["Off","Warning","Alarm","Alarm or warning"]
                            return(list.index(Affecter_niveau_diagnostic))
                        
                        def traduction_capt_to_ordi_Mode_défaut(Mode_défaut):
                            list=["NaN value","Last valid value","Alarm","Alarm or warning"]
                            return(list[Mode_défaut])  

                        def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                            list=["NaN value","Last valid value","Alarm","Alarm or warning"]
                            return(list.index(Mode_défaut))
                        
                        def traduction_capt_to_ordi_Mode_interpréteur (Mode_interpréteur ):
                            list=["Standard","Ignore surplus bytes"]
                            return(list[Mode_interpréteur ])  

                        def traduction_ordi_to_capt_Mode_interpréteur(Mode_interpréteur):
                            list=["Standard","Ignore surplus bytes"]
                            return(list.index(Mode_interpréteur))
                
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t Communication")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Configuration Modbus")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Configuration")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Adresse Bus",recuperationdonne.Adresse_Bus,4909)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Baudrate",traduction_capt_to_ordi_Baudrate(recuperationdonne.Baudrate),["1200 BAUD","2400 BAUD","4800 BAUD""9600 BAUD","19200 BAUD","38400 BAUD","57600 BAUD","115200 BAUD"],4911)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode de transfert de données     ",traduction_capt_to_ordi_Mode_de_transfert_de_données(recuperationdonne.Mode_de_transfert_de_données),["RTU","ASCII"],4912)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Parité",traduction_capt_to_ordi_Parité(recuperationdonne.Parité),["Even","Odd","None / 2 stop bits","None / 1 stop bit"],4913)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Ordre des octets",traduction_capt_to_ordi_Ordre_des_octets(recuperationdonne.Byte_order),["0-1-2-3","3-2-1-0","2-3-0-1","1-0-3-2"],4914)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Délai Télégramme",recuperationdonne.Telegram_delay,4915,1)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Affecter niveau diagnostic   ",traduction_capt_to_ordi_Affecter_niveau_diagnostic(recuperationdonne.Assign_diagnostic_behavior),["Off","Warning","Alarm","Alarm or warning"],4921)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode défaut",traduction_capt_to_ordi_Mode_défaut(recuperationdonne.Failure_mode),["NaN value","Last valid value","Alarm","Alarm or warning"],4920)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode interpréteur    ",traduction_capt_to_ordi_Mode_interpréteur(recuperationdonne.Interpreter_mode),["Standard","Ignore surplus bytes"],4924)

                        def traduction_capt_to_ordi_RAZ_tous_les_totalisateurs(RAZ_tous_les_totalisateurs ):
                            list=["Cancel","Reset + totalize"]
                            return(list[RAZ_tous_les_totalisateurs ])  

                        def traduction_ordi_to_capt_RAZ_tous_les_totalisateurs(RAZ_tous_les_totalisateurs):
                            list=["Cancel","Reset + totalize"]
                            return(list.index(RAZ_tous_les_totalisateurs))


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t Application")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Application")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t RAZ tous les totalisateurs",traduction_capt_to_ordi_RAZ_tous_les_totalisateurs(recuperationdonne.RAZ_tous_les_totalisateurs),["Cancel","Reset + totalize"],2608)

                        def traduction_capt_to_ordi_Affecter_variable_process_1(Affecter_variable_process_1 ):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list[Affecter_variable_process_1 ])  

                        def traduction_ordi_to_capt_Affecter_variable_process_1(Affecter_variable_process_1):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list.index(Affecter_variable_process_1))
                        
                        def traduction_capt_to_ordi_Unité_de_masse(Unité_de_masse ):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list[Unité_de_masse ])  

                        def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list.index(Unité_de_masse))
                        
                        def traduction_capt_to_ordi_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur ):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list[Mode_de_fonctionnement_totalisateur ])  

                        def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list.index(Mode_de_fonctionnement_totalisateur))
                        
                        def traduction_capt_to_ordi_Contrôle_totalisateur_1(Contrôle_totalisateur_1 ):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list[Contrôle_totalisateur_1 ])  

                        def traduction_ordi_to_capt_Contrôle_totalisateur_1(Contrôle_totalisateur_1):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list.index(Contrôle_totalisateur_1))
                        

                        def traduction_capt_to_ordi_Mode_défaut(Mode_défaut ):
                            list=["Stop","Actual value","Last valid value"]
                            return(list[Mode_défaut ])  

                        def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                            list=["Stop","Actual value","Last valid value"]
                            return(list.index(Mode_défaut))

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Totalisateur 1")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Totalisateur 1")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Affecter variable process    ",traduction_capt_to_ordi_Affecter_variable_process_1(recuperationdonne.Affecter_variable_process_1),['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow'],2600)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de masse",traduction_capt_to_ordi_Unité_de_masse(recuperationdonne.Unité_de_masse_1),['g','kg','t','oz','lb','STon','User mass'],2601)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode de fonctionnement totalisateur",traduction_capt_to_ordi_Mode_de_fonctionnement_totalisateur(recuperationdonne.Mode_de_fonctionnement_totalisateur_1),['Net flow total (Default)','Forward flow total','Reverse flow total'],2604 )
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Contrôle totalisateur 1",traduction_capt_to_ordi_Contrôle_totalisateur_1(recuperationdonne.Contrôle_totalisateur_1),["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"],2607)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur de présélection 1",(recuperationdonne.Valeur_de_présélection_1),2589,1)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode défaut",traduction_capt_to_ordi_Mode_défaut(recuperationdonne.Mode_défaut_1),["Stop","Actual value","Last valid value"],2605)

                        def traduction_capt_to_ordi_Affecter_variable_process_2(Affecter_variable_process_2 ):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list[Affecter_variable_process_2 ])  

                        def traduction_ordi_to_capt_Affecter_variable_process_2(Affecter_variable_process_2):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list.index(Affecter_variable_process_2))
                        
                        def traduction_capt_to_ordi_Unité_de_masse(Unité_de_masse ):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list[Unité_de_masse ])  

                        def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list.index(Unité_de_masse))
                        
                        def traduction_capt_to_ordi_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur ):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list[Mode_de_fonctionnement_totalisateur ])  

                        def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list.index(Mode_de_fonctionnement_totalisateur))
                        
                        def traduction_capt_to_ordi_Contrôle_totalisateur_2(Contrôle_totalisateur_2 ):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list[Contrôle_totalisateur_2 ])  

                        def traduction_ordi_to_capt_Contrôle_totalisateur_2(Contrôle_totalisateur_2):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list.index(Contrôle_totalisateur_2))

                        def traduction_capt_to_ordi_Mode_défaut(Mode_défaut ):
                            list=["Stop","Actual value","Last valid value"]
                            return(list[Mode_défaut ])  

                        def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                            list=["Stop","Actual value","Last valid value"]
                            return(list.index(Mode_défaut))
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Totalisateur 2")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Totalisateur 2")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Affecter variable process ",traduction_capt_to_ordi_Affecter_variable_process_2(recuperationdonne.Affecter_variable_process_2),['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow'],2800)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de masse",traduction_capt_to_ordi_Unité_de_masse(recuperationdonne.Unité_de_masse_2),['g','kg','t','oz','lb','STon','User mass'],2801)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode de fonctionnement totalisateur",traduction_capt_to_ordi_Mode_de_fonctionnement_totalisateur(recuperationdonne.Mode_de_fonctionnement_totalisateur_2),['Net flow total (Default)','Forward flow total','Reverse flow total'],2802 )
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Contrôle totalisateur 1",traduction_capt_to_ordi_Contrôle_totalisateur_2(recuperationdonne.Contrôle_totalisateur_2),["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"],2807)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur de présélection 1",(recuperationdonne.Valeur_de_présélection_2),2591,1)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode défaut",traduction_capt_to_ordi_Mode_défaut(recuperationdonne.Mode_défaut_2),["Stop","Actual value","Last valid value"],2805)

                        def traduction_capt_to_ordi_Affecter_variable_process_3(Affecter_variable_process_3 ):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list[Affecter_variable_process_3 ])  

                        def traduction_ordi_to_capt_Affecter_variable_process_3(Affecter_variable_process_3):
                            list=['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow']
                            return(list.index(Affecter_variable_process_3))
                        
                        def traduction_capt_to_ordi_Unité_de_masse(Unité_de_masse ):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list[Unité_de_masse ])  

                        def traduction_ordi_to_capt_Unité_de_masse(Unité_de_masse):
                            list=['g','kg','t','oz','lb','STon','User mass']
                            return(list.index(Unité_de_masse))
                        
                        def traduction_capt_to_ordi_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur ):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list[Mode_de_fonctionnement_totalisateur ])  

                        def traduction_ordi_to_capt_Mode_de_fonctionnement_totalisateur(Mode_de_fonctionnement_totalisateur):
                            list=['Net flow total (Default)','Forward flow total','Reverse flow total']
                            return(list.index(Mode_de_fonctionnement_totalisateur))
                        
                        def traduction_capt_to_ordi_Contrôle_totalisateur_3(Contrôle_totalisateur_3 ):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list[Contrôle_totalisateur_3 ])  

                        def traduction_ordi_to_capt_Contrôle_totalisateur_3(Contrôle_totalisateur_3):
                            list=["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"]
                            return(list.index(Contrôle_totalisateur_3))

                        def traduction_capt_to_ordi_Mode_défaut(Mode_défaut ):
                            list=["Stop","Actual value","Last valid value"]
                            return(list[Mode_défaut ])  

                        def traduction_ordi_to_capt_Mode_défaut(Mode_défaut):
                            list=["Stop","Actual value","Last valid value"]
                            return(list.index(Mode_défaut))


                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Totalisateur 3")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Totalisateur 3")
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Affecter variable process    ",traduction_capt_to_ordi_Affecter_variable_process_3(recuperationdonne.Affecter_variable_process_3),['Off','Mass flow (Default)','Volume flow','Corrected volume flow','Target mass flow','Carrier mass flow'],3000)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Unité de masse",traduction_capt_to_ordi_Unité_de_masse(recuperationdonne.Unité_de_masse_3),['g','kg','t','oz','lb','STon','User mass'],3001)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode de fonctionnement totalisateur",traduction_capt_to_ordi_Unité_de_masse(recuperationdonne.Mode_de_fonctionnement_totalisateur_3),['Net flow total (Default)','Forward flow total','Reverse flow total'],3004 )
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Contrôle totalisateur 1",traduction_capt_to_ordi_Contrôle_totalisateur_3(recuperationdonne.Contrôle_totalisateur_3),["Totalize","Reset + totalize","Preset + hold","Reset + hold","Preset + totalize"],2607)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t Valeur de présélection 1",(recuperationdonne.Valeur_de_présélection_3),2589,1)
                        self.scrollable_label_button_frame.add_liste_a_choix("\t\t\t\t\t\t Mode défaut",traduction_capt_to_ordi_Mode_défaut(recuperationdonne.Mode_défaut_3),["Stop","Actual value","Last valid value"],2605)

                        #page 13
                        """
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t Diagnostic")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Diagnostic")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Diagnostic actuel",recuperationdonne.Diagnostic_actuel,2731)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Dernier diagnostic",recuperationdonne.Dernier_diagnostic,2718)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Temps de fct depuis redémarrage",recuperationdonne.Temps_de_fct_depuis_redémarrage,2623)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t Temps de fonctionnement",recuperationdonne.Temps_de_fonctionnement,2630)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Diagnostic actuel")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Horodatage",recuperationdonne.)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Diagnostic diagnostic")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Horodatage",recuperationdonne.)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Liste de diagnostic")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Liste de diagnostic")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Diagnostic 1",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Diagnostic 2",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Diagnostic 3",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Diagnostic 4",recuperationdonne.)
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Diagnostic 5",recuperationdonne.)
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Diagnostic 1")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Horodatage",recuperationdonne.)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Diagnostic 2")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Horodatage","value")

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Diagnostic 3")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Horodatage","value")

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Diagnostic 4")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Horodatage","value")

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Diagnostic 5")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Horodatage","value")
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Journal d'événements")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Journal d'événements")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Options filtre","value")
                    
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Information appareil")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Information appareil")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Désignation du point de mesure","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Numéro de série","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Version logiciel","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Nom d'appareil","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Code commande","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Version ENP","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Compteur configuration","value")
                        
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Référence de commande")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Référence de commande 1","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Référence de commande 2","value")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Référence de commande 3","value")
                        """
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Valeurs min. / max.")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Température électronique")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Température électronique")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Température_électronique) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Température_électronique)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Température du fluide")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Température du fluide")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Température_du_fluide) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Température_du_fluide)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Fréquence d'oscillation")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Fréquence d'oscillation")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Fréquence_d_oscillation) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Fréquence_d_oscillation)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Amplitude de l'oscillation")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Amplitude de l'oscillation")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Amplitude_de_l_oscillation) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Amplitude_de_l_oscillation)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Amplitude de l'oscillation")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Amplitude de l'oscillation")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Amortissement_de_l_oscillation) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Amortissement_de_l_oscillation)

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Asymétrie signal")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t\t Asymétrie signal")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur minimale",recuperationdonne.Valeur_minimale_Asymétrie_signal) 
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t\t Valeur maximale",recuperationdonne.Valeur_maximale_Asymétrie_signal)
                        """"
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t Simulation")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Simulation 1")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Affecter simulation variable process",recuperationdonne.) 

                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t Simulation 3")
                        self.scrollable_label_button_frame.add_item(f"\t\t\t\t\t\t\t\t Simulation alarme apparei",recuperationdonne.) 
                        """
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t")
                        self.scrollable_label_button_frame.add_itemtitle(f"\t\t\t\t\t\t\t")
                        

                        attempt += 1
                except filelock.Timeout:
                    attempt += 1
                    print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
                    if attempt < max_attempts:
                        time.sleep(1)  # Optional: Add a delay before retrying

                if attempt == max_attempts:
                    print("Maximum attempts reached. Lock could not be acquired.")
        lock_excel_file()
