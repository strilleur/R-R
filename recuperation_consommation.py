import openpyxl
import time
from filelock import Timeout, FileLock
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import time
import customtkinter
from datetime import datetime
from datetime import time
import minimalmodbus 
file_path_Etat_Communication = 'fichier_excel/Etat Communication.xlsx'
file_path_consommation = 'fichier_excel/consommation.xlsx'
sheet_name = "Feuille1"



import openpyxl
import filelock
import time

dictconsommation= {}
list_time=[datetime.now()]

def consommation():
    # Create a file lock
    file_lock = filelock.FileLock(file_path_Etat_Communication + ".lock")
    max_attempts = 1  # Maximum attempts to acquire the lock
    attempt = 0
    
    while attempt < max_attempts:
        try:
            # Acquire the lock with a timeout of 20 seconds
            with file_lock.acquire(timeout=20):
                print(f"Lock acquired for first {file_path_Etat_Communication}")
                
                # Perform operations on the Excel file
                wb = openpyxl.load_workbook(file_path_Etat_Communication)
                sheet = wb.active

                def update():

                    """
                    On va check si Le BUS est en cour de communication.
                    Si oui on va attendre 0.5sec et check a nouveau
                    Si non on  va lancé le programme principale 
                    """
                    """
                    Le programme principale:
                    Tout d'abord on va importer du module configuration les liens adress-nom des capteurs.
                    Puis on import du module recuperation_donne les consommation qu'on va mettre dans un dictionnaire.

                    """
                    
                    
                    try:
                        from configuration import window_configuration
                    except:
                        print("error importation configuration")

                    
                    
                    window_configuration=window_configuration("admin")
                    
                    
                    liste_d_address=window_configuration.listconfig
                    window_configuration.destroy()
                    try:
                        print("try recuperation_donne")
                        from recuperation_donne import recuperationdonne
                    except:
                        print("error importation recuperation_donne")
                    
                    for cle,valeurs in liste_d_address.items():
                        if valeurs!="None":
                            true_valeurs=int(valeurs)
                            
                            rs485communication = minimalmodbus.Instrument('/dev/ttyUSB0',true_valeurs)	# Make an "instrument" object called rs485communication (port name, slave address (in decimal))
                            rs485communication.serial.baudrate = 19200				# BaudRate
                            rs485communication.serial.bytesize = 8					# Number of data bits to be requested
                            rs485communication.serial.parity = minimalmodbus.serial.PARITY_EVEN	# Parity Setting here is NONE but can be ODD or EVEN
                            rs485communication.serial.stopbits = 1					# Number of stop bits
                            rs485communication.serial.timeout  = 0.5					# Timeout time in seconds
                            rs485communication.mode = minimalmodbus.MODE_RTU		# Mode to be used (RTU or ascii mode)
                            try:
                                consommation=rs485communication.read_float(2609,3,2,3)
                                dictconsommation[cle]=str(consommation)
                            except:
                                dictconsommation[cle]=str("0")

                            
                        else:
                            dictconsommation[cle]=str(0)
                attempt += 1
                update()
                
        except filelock.Timeout:
            attempt += 1
            print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
            if attempt < max_attempts:
                time.sleep(1)  # Optional: Add a delay before retrying

        if attempt == max_attempts:
            pass
                    
    file_lock = filelock.FileLock(file_path_consommation + ".lock")

    max_attempts_consommation = 1  # Maximum attempts to acquire the lock
    attempt_consommation = 0
          
    while attempt_consommation < max_attempts_consommation:
        
        try:
            # Acquire the lock with a timeout of 20 seconds
            with file_lock.acquire(timeout=20):
                print(f"Lock acquired for {file_path_consommation}")
                                
                # Perform operations on the Excel file
                workbook = openpyxl.load_workbook(file_path_consommation)
                worksheet = workbook.active

                def update_consommation():
                    from datetime import datetime
                    from datetime import time
                    from datetime import timedelta
                    
                    
                    dt = datetime.now()
                    enter_delta= list_time.pop()
                    exit_delta=dt
                    
                    
                    difference_delta = exit_delta - enter_delta
                    
                    timedelta_1m = timedelta(days=0,seconds=59,microseconds=0,milliseconds=0,minutes=4,hours=0,weeks=0)
                    timedelta_10s = timedelta(days=0,seconds=9,microseconds=0,milliseconds=0,minutes=0,hours=0,weeks=0)
                    
                    if difference_delta>=timedelta_1m:
                        list_time.append(dt)
                        x = dt.strftime("%Y-%m-%d %H:%M:%S")
                        worksheet.append({'A': x ,'C': dictconsommation["MP1IN"], 'D':dictconsommation["MP1OUT"], 'E': str(float(dictconsommation["MP1OUT"])-float(dictconsommation["MP1IN"])), 'F': dictconsommation["MP2IN"], 'G':dictconsommation["MP2OUT"],'H': str(float(dictconsommation["MP2OUT"])-float(dictconsommation["MP2IN"])), 'I': dictconsommation["MP3IN"], 'J': dictconsommation["MP3OUT"], 'K': str(float(dictconsommation["MP3OUT"])-float(dictconsommation["MP3IN"])),'L': dictconsommation["MP4IN"], 'M': dictconsommation["MP4OUT"], 'N': str(float(dictconsommation["MP4OUT"])-float(dictconsommation["MP4IN"])), 'O': dictconsommation["GE1IN"],'P': dictconsommation["GE1OUT"], 'Q': str(float(dictconsommation["GE1OUT"])-float(dictconsommation["GE1IN"])) , 'R':dictconsommation["GE2IN"] , 'S':dictconsommation["GE2OUT"] ,'T': str(float(dictconsommation["GE2OUT"])-float(dictconsommation["GE2IN"])), 'U': dictconsommation["GE3IN"], 'V': dictconsommation["GE3OUT"], 'W': str(float(dictconsommation["GE3OUT"])-float(dictconsommation["GE3IN"])),'X': dictconsommation["GE4IN"], 'Y': dictconsommation["GE4OUT"],'Z': str(float(dictconsommation["GE4OUT"])-float(dictconsommation["GE4IN"]))})
                        x = dt.strftime("%Y-%m-%d %H:%M:%S")
                        worksheet.delete_cols(27,53)
                        worksheet.append({'AA': x , 'AC': dictconsommation["MP1IN"], 'AD':dictconsommation["MP1OUT"], 'AE': str(float(dictconsommation["MP1OUT"])-float(dictconsommation["MP1IN"])), 'AF': dictconsommation["MP2IN"], 'AG':dictconsommation["MP2OUT"],'AH': str(float(dictconsommation["MP2OUT"])-float(dictconsommation["MP2IN"])), 'AI': dictconsommation["MP3IN"], 'AJ': dictconsommation["MP3OUT"], 'AK': str(float(dictconsommation["MP3OUT"])-float(dictconsommation["MP3IN"])),'AL': dictconsommation["MP4IN"], 'AM': dictconsommation["MP4OUT"], 'AN': str(float(dictconsommation["MP4OUT"])-float(dictconsommation["MP4IN"])), 'AO': dictconsommation["GE1IN"],'AP': dictconsommation["GE1OUT"], 'AQ': str(float(dictconsommation["GE1OUT"])-float(dictconsommation["GE1IN"])) , 'AR':dictconsommation["GE2IN"] , 'AS':dictconsommation["GE2OUT"] ,'AT': str(float(dictconsommation["GE2OUT"])-float(dictconsommation["GE2IN"])), 'AU': dictconsommation["GE3IN"], 'AV': dictconsommation["GE3OUT"], 'AW': str(float(dictconsommation["GE3OUT"])-float(dictconsommation["GE3IN"])),'AX': dictconsommation["GE4IN"], 'AY': dictconsommation["GE4OUT"],'AZ': str(float(dictconsommation["GE4OUT"])-float(dictconsommation["GE4IN"]))})
                    else:
                        x = dt.strftime("%Y-%m-%d %H:%M:%S")
                        
                        worksheet.delete_cols(27,53)
                        worksheet['AA1'] = x
                        worksheet['AC1'] = dictconsommation["MP1IN"]
                        worksheet['AD1'] = dictconsommation["MP1OUT"]
                        worksheet['AE1'] = str(float(dictconsommation["MP1IN"])-float(dictconsommation["MP1OUT"]))
                        worksheet['AF1'] = dictconsommation["MP2IN"]
                        worksheet['AG1'] = dictconsommation["MP2OUT"]
                        worksheet['AH1'] = str(float(dictconsommation["MP2IN"])-float(dictconsommation["MP2OUT"]))
                        worksheet['AI1'] = dictconsommation["MP3IN"]
                        worksheet['AJ1'] = dictconsommation["MP3OUT"]
                        worksheet['AK1'] = str(float(dictconsommation["MP3IN"])-float(dictconsommation["MP3OUT"]))
                        worksheet['AL1'] = dictconsommation["MP4IN"]
                        worksheet['AM1'] = dictconsommation["MP4OUT"]
                        worksheet['AN1'] = str(float(dictconsommation["MP4IN"])-float(dictconsommation["MP4OUT"]))
                        worksheet['AO1'] = dictconsommation["GE1IN"]
                        worksheet['AP1'] = dictconsommation["GE1OUT"]
                        worksheet['AQ1'] = str(float(dictconsommation["GE1IN"])-float(dictconsommation["GE1OUT"]))
                        worksheet['AR1'] = dictconsommation["GE2IN"]
                        worksheet['AS1'] = dictconsommation["GE2OUT"]
                        worksheet['AT1'] = str(float(dictconsommation["GE2IN"])-float(dictconsommation["GE2OUT"]))
                        worksheet['AU1'] = dictconsommation["GE3IN"]
                        worksheet['AV1'] = dictconsommation["GE3OUT"]
                        worksheet['AW1'] = str(float(dictconsommation["GE3IN"])-float(dictconsommation["GE3OUT"]))
                        worksheet['AX1'] = dictconsommation["GE4IN"]
                        worksheet['AY1'] = dictconsommation["GE4OUT"]
                        worksheet['AZ1'] = str(float(dictconsommation["GE4IN"])-float(dictconsommation["GE4OUT"]))
                        list_time.append(enter_delta)
                        

                    workbook.save("fichier_excel/consommation.xlsx")
                    workbook.close()
                    
                    
                update_consommation()
                attempt_consommation += 1        
        except filelock.Timeout:
            attempt_consommation += 1
            print(f"Attempt {attempt_consommation}: Lock acquisition timed out. Retrying...")
            if attempt_consommation < max_attempts_consommation:
                time.sleep(1)  # Optional: Add a delay before retrying
            # Exit the loop if the lock is acquired successfully
            attempt=attempt+1
    
if __name__ == "__main__":
    while True:
        dt1 = datetime.now()
        consommation()
        dt2 = datetime.now()
        Dt=dt2-dt1
        time_for_excecution=Dt.total_seconds()
        if time_for_excecution>10:
            pass
        else:
            time_to_wait=10-time_for_excecution
            time.sleep(time_to_wait)

    

