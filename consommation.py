import tkinter
#import tkintermapview
import customtkinter
import os
from PIL import Image
from openpyxl import Workbook 
from openpyxl import load_workbook
import datetime
import subprocess
from datetime import datetime
import time
import openpyxl
import filelock
import subprocess

max_attempts = 1  # Maximum attempts to acquire the lock
attempt = 0

traversé_quai_found=0
file_path_consommation = 'fichier_excel/consommation.xlsx'

workbook2 = openpyxl.load_workbook(file_path_consommation)

worksheet = workbook2.active
file_lock = filelock.FileLock(file_path_consommation + ".lock")
while attempt < max_attempts:
    try:
        # Acquire the lock with a timeout of 20 seconds
            with file_lock.acquire(timeout=20):
                print(f"Lock acquired for first {file_path_consommation}")
                        
                rowmax=worksheet.max_row
                MP1depart=str(worksheet.cell(row=(rowmax)+1, column=5).value)
                print(MP1depart)
                attempt += 1
    except filelock.Timeout:
        attempt += 1
        print(f"Attempt {attempt}: Lock acquisition timed out. Retrying...")
    if attempt < max_attempts:
        time.sleep(1)  # Optional: Add a delay before retrying

    if attempt == max_attempts:
        print("Maximum attempts reached. Lock could not be acquired.")
        break